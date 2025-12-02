import streamlit as st
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
import io
import os
import glob
import re
import unicodedata
from datetime import datetime
from urllib.parse import quote
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from fpdf import FPDF
import numpy as np # Necesario para algunas funciones de cálculo

# --- CONFIGURACIÓN VISUAL PROFESIONAL ---
st.set_page_config(
    page_title="Centro de Mando: Cobranza Ferreinox",
    page_icon="🛡️",
    layout="wide",
    initial_sidebar_state="expanded"
)

# Paleta de Colores y CSS Corporativo
COLOR_PRIMARIO = "#003366"  # Azul oscuro corporativo
COLOR_ACCION = "#FFC300"    # Amarillo para acciones
COLOR_FONDO = "#f4f6f9"
st.markdown(f"""
<style>
    .main {{ background-color: {COLOR_FONDO}; }}
    /* Métricas */
    .stMetric {{ background-color: white; padding: 15px; border-radius: 8px; border-left: 5px solid {COLOR_PRIMARIO}; box-shadow: 0 4px 6px rgba(0,0,0,0.1); }}
    /* Expander/Títulos de Sección */
    div[data-testid="stExpander"] div[role="button"] p {{ font-size: 1.1rem; font-weight: bold; color: {COLOR_PRIMARIO}; }}
    /* Título de la Aplicación */
    .css-1av0etd {{ color: {COLOR_PRIMARIO}; }} 
    /* Estilo de Botón de Acción (Simulación de botón de WhatsApp/PDF) */
    .action-button {{
        display: inline-block;
        padding: 8px 15px;
        color: white;
        background-color: #25D366; /* Verde WhatsApp */
        border-radius: 5px;
        text-align: center;
        text-decoration: none;
        font-weight: bold;
        transition: background-color 0.3s;
    }}
    .action-button:hover {{ background-color: #128C7E; }}
</style>
""", unsafe_allow_html=True)

# ======================================================================================
# 1. MOTOR DE INGESTIÓN Y LIMPIEZA DE DATOS (Inteligencia de Columnas)
# ======================================================================================

def normalizar_texto(texto):
    """Elimina tildes, símbolos y pone mayúsculas para mapeo."""
    if not isinstance(texto, str): return str(texto)
    texto = unicodedata.normalize('NFD', texto).encode('ascii', 'ignore').decode("utf-8").upper().strip()
    # Eliminar símbolos que no sean esenciales para el mapeo
    return re.sub(r'[^\w\s\.]', '', texto).strip()

def limpiar_moneda(valor):
    """Limpia formatos de moneda, tolerante a comas y puntos."""
    if pd.isna(valor): return 0.0
    s_val = str(valor).strip()
    s_val = re.sub(r'[^\d.,-]', '', s_val)
    if not s_val: return 0.0
    try:
        # Intenta manejar formatos (1.000,00 vs 1,000.00)
        if s_val.count(',') > 1 and s_val.count('.') == 0: s_val = s_val.replace(',', '') # Miles con coma
        elif s_val.count('.') > 1 and s_val.count(',') == 0: s_val = s_val.replace('.', '') # Miles con punto
        elif s_val.count(',') == 1 and s_val.count('.') == 1:
            if s_val.rfind(',') > s_val.rfind('.'): s_val = s_val.replace('.', '').replace(',', '.') # Latino (último es separador decimal)
            else: s_val = s_val.replace(',', '') # USA (coma es separador de miles)
        elif s_val.count(',') == 1 and s_val.count('.') == 0 and len(s_val.split(',')[-1]) <= 2: # Asume decimal con coma
            s_val = s_val.replace(',', '.')
        
        # Elimina el separador de miles si es solo un punto o coma (ya que el último se manejó)
        s_val = s_val.replace(',', '').replace('.', '') # Se eliminan para prevenir errores
        return float(s_val.replace(',', '').replace(' ', ''))
    except:
        try:
             # Caso final para forzar la conversión
            return float(re.sub(r'[^\d.]', '', s_val))
        except:
             return 0.0

def mapear_y_limpiar_df(df):
    """Mapea, limpia y valida las columnas."""
    df.columns = [normalizar_texto(c) for c in df.columns]
    
    # Mapeo de Columnas Críticas y Opcionales
    mapa = {
        'cliente': ['NOMBRE', 'RAZON SOCIAL', 'TERCERO', 'CLIENTE'],
        'nit': ['NIT', 'IDENTIFICACION', 'CEDULA', 'RUT'],
        'saldo': ['IMPORTE', 'SALDO', 'TOTAL', 'DEUDA', 'VALOR'],
        'dias': ['DIAS', 'VENCIDO', 'MORA', 'ANTIGUEDAD'],
        'telefono': ['TEL', 'MOVIL', 'CELULAR', 'TELEFONO', 'CONTACTO', 'TELF'],
        'vendedor': ['VENDEDOR', 'ASESOR', 'COMERCIAL', 'NOMVENDEDOR'],
        'factura': ['NUMERO', 'FACTURA', 'DOC', 'SERIE'],
        'email': ['CORREO', 'EMAIL', 'E-MAIL', 'MAIL']
    }
    
    renombres = {}
    for standard, variantes in mapa.items():
        for col in df.columns:
            # Una lógica más flexible que busca coincidencias parciales y exactas
            col_norm = normalizar_texto(col)
            if standard not in renombres.values() and any(v in col_norm for v in variantes):
                renombres[col] = standard
                break
    
    df.rename(columns=renombres, inplace=True)
    
    # --- VALIDACIÓN CRÍTICA ---
    req = ['cliente', 'saldo', 'dias']
    if not all(c in df.columns for c in req):
        missing = [c for c in req if c not in df.columns]
        return None, f"Faltan columnas críticas: {', '.join(missing)}. Columnas detectadas: {list(df.columns)}"

    # --- LIMPIEZA Y CONVERSIÓN ---
    df['saldo'] = df['saldo'].apply(limpiar_moneda)
    df['dias'] = pd.to_numeric(df['dias'], errors='coerce').fillna(0).astype(int)
    
    # Asegurar campos opcionales
    for c in ['telefono', 'vendedor', 'nit', 'factura', 'email']:
        if c not in df.columns: 
            df[c] = 'N/A'
        else:
            df[c] = df[c].fillna('N/A').astype(str)

    # Filtrar saldos > 0 y eliminar duplicados (ej. por factura)
    df_filtrado = df[df['saldo'] > 0]
    
    # Si hay columna de 'factura', se considera cada una como una deuda única
    if 'factura' in df_filtrado.columns and 'cliente' in df_filtrado.columns:
        df_filtrado.drop_duplicates(subset=['cliente', 'factura'], keep='first', inplace=True)
    
    return df_filtrado, "Datos limpios y listos."


@st.cache_data(ttl=600)
def cargar_datos(uploaded_file):
    """Procesa el archivo subido por el usuario."""
    if uploaded_file is None:
        return None, "Por favor, sube un archivo de Cartera."

    try:
        # Detección del tipo de archivo
        if uploaded_file.name.endswith('.csv'):
            df = pd.read_csv(uploaded_file, sep=None, engine='python', encoding='latin-1', dtype=str)
        else:
            df = pd.read_excel(uploaded_file, engine='openpyxl', dtype=str)
            
        df_procesado, status = mapear_y_limpiar_df(df)
        
        if df_procesado is None:
            return None, f"Error en la estructura del archivo {uploaded_file.name}: {status}"
        
        return df_procesado, f"Datos cargados y limpios de: {uploaded_file.name}"
        
    except Exception as e:
        return None, f"Error leyendo {uploaded_file.name}: {str(e)}"

# ======================================================================================
# 2. CEREBRO DE ESTRATEGIA, KPIs y Gestión
# ======================================================================================

def generar_estrategia(df):
    """Segmenta la cartera en rangos de edad y asigna prioridad."""
    
    bins = [-float('inf'), 0, 15, 30, 60, float('inf')]
    labels = ["🟢 Corriente (0)", "🟡 Preventivo (1-15)", "🟠 Administrativo (16-30)", "🔴 Alto Riesgo (31-60)", "⚫ Pre-Jurídico (+60)"]
    prioridad_map = {"🟢 Corriente (0)": 5, "🟡 Preventivo (1-15)": 4, "🟠 Administrativo (16-30)": 3, "🔴 Alto Riesgo (31-60)": 2, "⚫ Pre-Jurídico (+60)": 1}
    
    df['Estado'] = pd.cut(df['dias'], bins=bins, labels=labels, right=True, ordered=True)
    df['Prioridad'] = df['Estado'].map(prioridad_map)
    
    return df

def calcular_kpis(df):
    """Calcula indicadores clave de rendimiento (KPIs)."""
    total = df['saldo'].sum()
    vencido = df[df['dias'] > 0]['saldo'].sum()
    critico_60_mas = df[df['dias'] >= 60]['saldo'].sum()
    
    pct_mora = (vencido / total) * 100 if total > 0 else 0
    
    # Cálculo del Índice de Severidad de Cobranza (CSI)
    # CSI = Suma(Saldo * Días Mora) / Saldo Total
    csi_numerator = (df['saldo'] * df['dias']).sum()
    csi = csi_numerator / total if total > 0 else 0
    
    # Antigüedad Promedio Vencida
    vencido_df = df[df['dias'] > 0]
    antiguedad_prom_vencida = (vencido_df['saldo'] * vencido_df['dias']).sum() / vencido if vencido > 0 else 0
    
    clientes_morosos = df[df['dias'] > 0]['cliente'].nunique()
    
    return {
        'total': total,
        'vencido': vencido,
        'critico_60_mas': critico_60_mas,
        'pct_mora': pct_mora,
        'csi': csi,
        'antiguedad_prom_vencida': antiguedad_prom_vencida,
        'clientes_morosos': clientes_morosos
    }

def crear_link_whatsapp(row, is_summary=False):
    """Genera el enlace de WhatsApp con un mensaje basado en el riesgo."""
    tel = str(row['telefono']).strip()
    tel = re.sub(r'\D', '', tel) # Quita todo lo que no sea dígito
    if len(tel) < 10: return None # Número incompleto
    if len(tel) == 10 and tel.startswith('3'): tel = '57' + tel # Asume Colombia si empieza con 3
    
    # En el modo de resumen por cliente, el saldo y los días son el total/máximo del cliente.
    saldo = row['Saldo_Total_Cliente'] if is_summary else row['saldo']
    dias = row['Dias_Max_Mora'] if is_summary else row['dias']
    cliente = str(row['cliente']).split()[0].title()
    
    # Lógica de Guion más agresiva y directa
    if dias <= 0:
        msg = f"Hola {cliente}, saludamos de Ferreinox. ¡Felicitaciones! Tu cuenta está al día. Hemos enviado tu estado de cuenta a tu correo. ¡Gracias por tu gestión!"
    elif dias <= 15:
        msg = f"Hola {cliente}, recordatorio preventivo de Ferreinox. Tienes un saldo pendiente de ${saldo:,.0f} (Factura: {row['factura'] if not is_summary else 'Varias'}) vencido hace {dias} días. ¿Puedes confirmar la fecha de pago HOY, por favor?"
    elif dias <= 30:
        msg = f"ATENCIÓN {cliente}: Su factura ${saldo:,.0f} ({row['factura'] if not is_summary else 'Varias'}) tiene {dias} días de vencimiento. Requerimos confirmación de pago. Si no recibimos respuesta hoy, su gestión pasará a un nivel superior."
    elif dias <= 60:
        msg = f"URGENTE {cliente}: Su deuda de ${saldo:,.0f} ({dias} días) pone su crédito bajo revisión. Debe realizar el pago antes de 24 horas para evitar el bloqueo total de despachos. Contáctenos inmediatamente."
    else:
        msg = f"ACCIÓN LEGAL {cliente}: Su cuenta está en estado PRE-JURÍDICO. Saldo: ${saldo:,.0f}. Evite reporte a centrales de riesgo y honorarios. *Responda HOY* con el comprobante de pago."
    
    if numero_limpio:
        return f"https://wa.me/{numero_limpio}?text={quote(msg)}"
    return None

# ======================================================================================
# 3. EXPORTACIÓN PROFESIONAL (Basado en el segundo código)
# ======================================================================================

def generar_excel_gerencial(df_input, kpis):
    """Genera un archivo Excel con formato corporativo, resumen ejecutivo y detalle."""
    output = io.BytesIO()
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Reporte Gerencial"
    
    # Estilos
    azul_corp = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    gris_claro = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
    rojo_alerta = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type='solid')
    header_font = Font(color="FFFFFF", bold=True)
    font_bold = Font(bold=True)
    money_fmt = '"$"#,##0'
    pct_fmt = '0.0%'
    
    thin_border = Border(left=Side(style='thin'), 
                         right=Side(style='thin'), 
                         top=Side(style='thin'), 
                         bottom=Side(style='thin'))

    # --- Resumen Ejecutivo (A1:F10) ---
    sheet['A1'] = "REPORTE EJECUTIVO DE CARTERA - FERREINOX"
    sheet['A1'].font = Font(size=18, bold=True, color="003366")
    sheet['A2'] = f"Generado: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    sheet['A2'].font = Font(italic=True)
    
    # KPIs en el resumen
    kpi_data = [
        ("TOTAL CARTERA", kpis['total'], money_fmt, azul_corp),
        ("TOTAL VENCIDO", kpis['vencido'], money_fmt, PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")),
        ("% DE MORA", kpis['pct_mora'] / 100, pct_fmt, gris_claro),
        ("CRÍTICO (+60 Días)", kpis['critico_60_mas'], money_fmt, rojo_alerta),
        ("ANTIGÜEDAD PROM. VENCIDA", kpis['antiguedad_prom_vencida'], '0 días', gris_claro),
        ("ÍNDICE DE SEVERIDAD (CSI)", kpis['csi'], '0.0', gris_claro),
    ]

    fila_kpi = 4
    for label, value, fmt, fill in kpi_data:
        sheet[f'A{fila_kpi}'] = label
        sheet[f'A{fila_kpi}'].font = font_bold
        sheet[f'A{fila_kpi}'].fill = fill if fila_kpi % 2 == 0 else gris_claro
        sheet[f'B{fila_kpi}'] = value
        sheet[f'B{fila_kpi}'].number_format = fmt
        sheet[f'B{fila_kpi}'].font = font_bold
        fila_kpi += 1

    # Separador para la tabla de detalle
    sheet['A11'] = "DETALLE COMPLETO DE CARTERA (Filtrable)"
    sheet['A11'].font = Font(size=14, bold=True, color="003366")
    
    # --- Detalle de Cartera (Fila 13 en adelante) ---
    
    # Columnas a Exportar
    cols = ['cliente', 'nit', 'factura', 'vendedor', 'telefono', 'email', 'Estado', 'dias', 'saldo']
    cols_to_export = [c for c in cols if c in df_input.columns]
    start_row = 13
    
    # Headers
    sheet.append([c.upper() for c in cols_to_export])
    for col_idx, cell in enumerate(sheet[start_row], 1):
        cell.fill = azul_corp
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        cell.border = thin_border
        
    # Rows
    for row in df_input[cols_to_export].itertuples(index=False):
        sheet.append(row)
        
    # Formato de Moneda y Colores por Días Vencidos
    saldo_col_idx = cols_to_export.index('saldo') + 1
    dias_col_idx = cols_to_export.index('dias') + 1
    
    for row_idx, row in enumerate(sheet.iter_rows(min_row=start_row + 1, max_row=sheet.max_row), start=start_row + 1):
        # Formato de saldo
        row[saldo_col_idx - 1].number_format = money_fmt
        
        # Colores por Mora
        dias = row[dias_col_idx - 1].value
        if isinstance(dias, int):
            if dias > 60: row[dias_col_idx - 1].fill = PatternFill(start_color="FF9999", end_color="FF9999", fill_type='solid') # Rojo Fuerte
            elif dias > 30: row[dias_col_idx - 1].fill = PatternFill(start_color="FFEB99", end_color="FFEB99", fill_type='solid') # Naranja
            elif dias > 0: row[dias_col_idx - 1].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type='solid') # Amarillo

    # Autoajuste de columnas
    for i, col_name in enumerate(cols_to_export, 1):
        ancho = 30 if col_name == 'cliente' else (20 if col_name in ['saldo', 'email', 'vendedor'] else 15)
        sheet.column_dimensions[chr(64 + i)].width = ancho
        
    # --- Configurar Filtros y Congelar Paneles ---
    ultima_fila_detalle = sheet.max_row
    sheet.auto_filter.ref = f"A{start_row}:{get_column_letter(len(cols_to_export))}{ultima_fila_detalle}"
    sheet.freeze_panes = f'A{start_row + 1}'

    workbook.save(output)
    return output.getvalue()


# --- CLASE PDF ESTADO DE CUENTA (Similar a la del segundo código) ---

class PDF(FPDF):
    def header(self):
        self.set_font('Arial', 'B', 18)
        self.set_text_color(0, 51, 102) # Azul Corporativo
        self.cell(0, 10, 'ESTADO DE CUENTA | FERREINOX SAS BIC', 0, 1, 'C')
        self.set_font('Arial', 'I', 9)
        self.set_text_color(100, 100, 100)
        self.cell(0, 5, f'Generado: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}', 0, 1, 'R')
        self.ln(5)
        self.set_draw_color(0, 51, 102)
        self.line(10, self.get_y(), 200, self.get_y())
        self.ln(5)

    def footer(self):
        self.set_y(-25)
        self.set_font('Arial', 'I', 9)
        self.set_text_color(100, 100, 100)
        self.cell(0, 5, "Portal de Pagos Ferreinox: [Enlace no incluido por seguridad]", 0, 1, 'C')
        self.set_y(-15)
        self.set_font('Arial', '', 8)
        self.cell(0, 10, f'Página {self.page_no()}/{{nb}}', 0, 0, 'C')

def generar_pdf_estado_cuenta(datos_cliente: pd.DataFrame, total_vencido_cliente: float):
    pdf = PDF()
    pdf.set_auto_page_break(auto=True, margin=30)
    pdf.add_page()
    pdf.alias_nb_pages()

    if datos_cliente.empty:
        pdf.set_font('Arial', 'B', 12)
        pdf.cell(0, 10, 'No se encontraron facturas para este cliente.', 0, 1, 'C')
        return bytes(pdf.output())

    datos_cliente_ordenados = datos_cliente.sort_values(by='dias', ascending=False)
    info_cliente = datos_cliente_ordenados.iloc[0]

    # Información del Cliente
    pdf.set_font('Arial', 'B', 11); pdf.cell(40, 10, 'Cliente:', 0, 0)
    pdf.set_font('Arial', '', 11); pdf.cell(0, 10, info_cliente['cliente'], 0, 1)
    pdf.set_font('Arial', 'B', 11); pdf.cell(40, 10, 'NIT:', 0, 0)
    pdf.set_font('Arial', '', 11); pdf.cell(0, 10, info_cliente['nit'], 0, 1); pdf.ln(5)

    # Mensaje de introducción
    mensaje = (f"Apreciado(a) {info_cliente['cliente'].split()[0].title()}, adjunto encontrará el detalle de su estado de cuenta a la fecha. "
               f"Le solicitamos su amable gestión para el pago de los valores vencidos.")
    pdf.set_text_color(50, 50, 50); pdf.set_font('Arial', 'I', 10); pdf.multi_cell(0, 5, mensaje, 0, 'J'); pdf.ln(5)
    
    # Encabezados de la tabla
    pdf.set_font('Arial', 'B', 9); pdf.set_fill_color(0, 51, 102); pdf.set_text_color(255, 255, 255)
    headers = [('Factura', 25), ('Días Mora', 20), ('Vendedor', 40), ('Saldo', 30)]
    for header, width in headers: pdf.cell(width, 8, header, 1, 0, 'C', 1)
    pdf.ln()

    # Filas de la tabla
    total_importe = 0
    for _, row in datos_cliente_ordenados.iterrows():
        total_importe += row['saldo']
        pdf.set_text_color(0, 0, 0)
        
        # Color de fondo según la mora
        if row['dias'] > 60: pdf.set_fill_color(255, 220, 220) # Rojo claro
        elif row['dias'] > 30: pdf.set_fill_color(255, 240, 220) # Naranja claro
        elif row['dias'] > 0: pdf.set_fill_color(255, 255, 220) # Amarillo claro
        else: pdf.set_fill_color(255, 255, 255) # Blanco para al día
        
        pdf.set_font('Arial', '', 9)
        pdf.cell(25, 6, str(row['factura']), 1, 0, 'C', 1)
        pdf.cell(20, 6, str(int(row['dias'])), 1, 0, 'C', 1)
        pdf.cell(40, 6, row['vendedor'], 1, 0, 'L', 1)
        pdf.cell(30, 6, f"${row['saldo']:,.0f}", 1, 0, 'R', 1)
        pdf.ln()

    # Fila de Totales
    pdf.set_text_color(0, 0, 0); pdf.set_fill_color(224, 224, 224); pdf.set_font('Arial', 'B', 9)
    pdf.cell(85, 8, 'TOTAL ADEUDADO', 1, 0, 'R', 1)
    pdf.set_fill_color(240, 240, 240)
    pdf.cell(30, 8, f"${total_importe:,.0f}", 1, 0, 'R', 1)
    pdf.ln()

    if total_vencido_cliente > 0:
        pdf.set_text_color(192, 0, 0); pdf.set_fill_color(255, 204, 204); pdf.set_font('Arial', 'B', 10)
        pdf.cell(85, 8, 'VALOR TOTAL VENCIDO', 1, 0, 'R', 1)
        pdf.cell(30, 8, f"${total_vencido_cliente:,.0f}", 1, 0, 'R', 1)
        pdf.ln()

    return bytes(pdf.output())

# ======================================================================================
# 4. INTERFAZ GRÁFICA (DASHBOARD)
# ======================================================================================

def main():
    col_logo, col_titulo, col_upload = st.columns([1, 4, 2])
    
    with col_titulo:
        st.title("🛡️ Centro de Mando: Cobranza Estratégica")
        st.markdown(f"**Ferreinox SAS BIC** | Panel Operativo y Gerencial | Última actualización: {datetime.now().strftime('%d/%m/%Y %H:%M')}")
        
    with col_upload:
        uploaded_file = st.file_uploader("📂 **Subir Archivo de Cartera** (`.xlsx` o `.csv`)", type=['xlsx', 'csv'], help="El archivo debe contener las columnas: Cliente, Saldo y Días Mora.")
        if uploaded_file:
            st.button("🔄 Recargar Datos (Con el archivo subido)", on_click=st.cache_data.clear)
            
    # --- CARGA Y PROCESAMIENTO ---
    df_raw, status = cargar_datos(uploaded_file)
    
    if df_raw is None:
        st.error(status)
        st.stop()

    # Aplicar Estrategia
    df = generar_estrategia(df_raw.copy())
    
    # --- SIDEBAR: FILTROS ---
    with st.sidebar:
        st.header("🔍 Filtros de Gestión")
        
        # Filtro Vendedor
        vendedores = ["TODOS"] + sorted(list(df['vendedor'].unique()))
        sel_vendedor = st.selectbox("Vendedor / Asesor", vendedores)
        if sel_vendedor != "TODOS":
            df = df[df['vendedor'] == sel_vendedor]

        # Filtro Estado (Prioridad)
        # Ordenar por el número de prioridad (cuanto menor, más crítico)
        estados_ordenados = sorted(list(df['Estado'].unique()), key=lambda x: df[df['Estado'] == x]['Prioridad'].mean(), reverse=False)
        estados = ["TODOS"] + estados_ordenados
        sel_estado = st.selectbox("Estado de Mora", estados)
        if sel_estado != "TODOS":
            df = df[df['Estado'] == sel_estado]
        
        st.markdown("---")
        mostrar_vencido = st.checkbox("Mostrar solo Cartera Vencida")
        if mostrar_vencido:
             df = df[df['dias'] > 0]
        st.markdown("---")
        st.markdown("### ⚙️ Herramienta de Gestión Rápida")
        # Generar DataFrame agrupado por cliente para la gestión individual
        df_clientes_gestion = df.groupby('cliente').agg(
            Saldo_Total_Cliente=('saldo', 'sum'),
            Dias_Max_Mora=('dias', 'max'),
            Vendedor=('vendedor', lambda x: x.iloc[0]),
            Telefono=('telefono', lambda x: x.iloc[0]),
            Email=('email', lambda x: x.iloc[0])
        ).reset_index()
        
        # Generar Link WA para el resumen
        df_clientes_gestion['Link_WA'] = df_clientes_gestion.apply(lambda row: crear_link_whatsapp(row, is_summary=True), axis=1)

        lista_clientes = [""] + sorted(df_clientes_gestion['cliente'].unique())
        sel_cliente_rapido = st.selectbox("🎯 Cliente para Contacto Directo", lista_clientes, index=0)

    if df.empty:
        st.warning("No hay datos que coincidan con los filtros aplicados.")
        return

    # --- CÁLCULO DE KPIS ---
    kpis = calcular_kpis(df)

    # --- KPIs SUPERIORES ---
    st.header("Indicadores Clave de Rendimiento (KPIs) 📈")
    k1, k2, k3, k4, k5, k6 = st.columns(6)
    
    k1.metric("💰 Cartera Total", f"${kpis['total']:,.0f}")
    k2.metric("⚠️ Total Vencido", f"${kpis['vencido']:,.0f}", f"{kpis['pct_mora']:.1f}%")
    k3.metric("🔥 Crítico (+60 Días)", f"${kpis['critico_60_mas']:,.0f}")
    k4.metric("👥 Clientes Morosos", f"{kpis['clientes_morosos']}")
    k5.metric("⏳ Antigüedad Prom. Vencida", f"{kpis['antiguedad_prom_vencida']:.0f} días")
    k6.metric("💥 Índice de Severidad (CSI)", f"{kpis['csi']:.1f}", help="Mayor CSI = Mayor riesgo de cuentas por cobrar. Suma(Saldo * Días) / Saldo Total")
    
    st.markdown("---")

    # --- PESTAÑAS PRINCIPALES ---
    tab_accion, tab_analisis, tab_export = st.tabs(["🚀 GESTIÓN DIARIA INMEDIATA", "📊 ANÁLISIS GERENCIAL", "📥 EXPORTAR Y DATOS"])

    # --------------------------------------------------------
    # TAB 1: GESTIÓN (Prioridad para Líder de Cartera)
    # --------------------------------------------------------
    with tab_accion:
        st.subheader("🎯 Tareas del Día: Cartera Vencida por Factura")
        st.caption("Ordenado por Prioridad (Pre-Jurídico a Preventivo) y Mayor Saldo. Actúe de arriba hacia abajo.")

        # Preparar datos para la tabla interactiva
        df_display = df[df['dias'] > 0].sort_values(by=['Prioridad', 'saldo'], ascending=[True, False]).copy()
        
        # Generar Link WA por factura (usando los datos de la fila individual)
        df_display['Link_WA'] = df_display.apply(lambda row: crear_link_whatsapp(row, is_summary=False), axis=1)
        
        columnas_accion = ['Estado', 'cliente', 'factura', 'dias', 'saldo', 'vendedor', 'telefono', 'Link_WA']
        
        st.data_editor(
            df_display[columnas_accion],
            column_config={
                "Link_WA": st.column_config.LinkColumn(
                    "📱 ACCIÓN WHATSAPP",
                    help="Clic para abrir WhatsApp Web con el guion listo",
                    validate="^https://wa\.me/.*",
                    display_text="💬 ENVIAR GUION"
                ),
                "saldo": st.column_config.NumberColumn("Deuda Factura", format="$ %d"),
                "dias": st.column_config.NumberColumn("Días Mora", format="%d días", min_value=1, max_value=365),
                "Estado": st.column_config.TextColumn("ESTADO (Prioridad)", width="medium"),
                "cliente": st.column_config.TextColumn("CLIENTE (Razón Social)", width="large"),
                "telefono": st.column_config.TextColumn("Teléfono")
            },
            hide_index=True,
            use_container_width=True,
            height=600
        )
        
        # Sección de Gestión Rápida Individual (para el cliente seleccionado en la sidebar)
        if sel_cliente_rapido:
            st.markdown("---")
            st.subheader(f"⚙️ Gestión Documental para: **{sel_cliente_rapido}**")
            info_cliente_rapida = df_clientes_gestion[df_clientes_gestion['cliente'] == sel_cliente_rapido].iloc[0]
            
            datos_cliente_detalle = df[df['cliente'] == sel_cliente_rapido]
            total_vencido_cliente = datos_cliente_detalle[datos_cliente_detalle['dias'] > 0]['saldo'].sum()
            
            pdf_bytes = generar_pdf_estado_cuenta(datos_cliente_detalle, total_vencido_cliente)
            
            col_pdf, col_wa_link = st.columns(2)
            
            with col_pdf:
                st.download_button(
                    label="📄 DESCARGAR ESTADO DE CUENTA (PDF)", 
                    data=pdf_bytes, 
                    file_name=f"Estado_Cuenta_{info_cliente_rapida['cliente'].replace(' ', '_')}_{datetime.now().strftime('%Y%m%d')}.pdf", 
                    mime="application/pdf"
                )
                st.info(f"Correo: {info_cliente_rapida['Email']} | Vendedor: {info_cliente_rapida['Vendedor']}")
            
            with col_wa_link:
                wa_url = info_cliente_rapida['Link_WA']
                if wa_url:
                    st.markdown(f'<a href="{wa_url}" target="_blank" class="action-button">📲 ENVIAR GUION RESUMEN A {info_cliente_rapida["Telefono"]}</a>', unsafe_allow_html=True)
                else:
                     st.warning("Número de WhatsApp no válido o no disponible.")


    # --------------------------------------------------------
    # TAB 2: ANÁLISIS (Visión Estratégica para Gerencia)
    # --------------------------------------------------------
    with tab_analisis:
        st.subheader("📈 Concentración y Antigüedad de la Cartera")
        c1, c2 = st.columns(2)
        
        # Gráfico de Distribución por Edad (Basado en el segundo código)
        with c1:
            st.markdown("### 1. Distribución de Cartera por Riesgo")
            df_edades = df.groupby('Estado', observed=True)['saldo'].sum().reset_index()
            color_map_edades = {"🟢 Corriente (0)": '#28A745', "🟡 Preventivo (1-15)": '#FFC107', "🟠 Administrativo (16-30)": '#FD7E14', "🔴 Alto Riesgo (31-60)": '#DC3545', "⚫ Pre-Jurídico (+60)": '#343A40'}
            fig_bar = px.bar(df_edades, x='Estado', y='saldo', text_auto='.2s', title='Monto de Cartera por Rango de Días', labels={'Estado': 'Antigüedad', 'saldo': 'Monto Total'}, color='Estado', color_discrete_map=color_map_edades)
            fig_bar.update_layout(xaxis={'categoryorder':'array', 'categoryarray': list(color_map_edades.keys())}, showlegend=False)
            st.plotly_chart(fig_bar, use_container_width=True)
            
        # Análisis de Pareto
        with c2:
            st.markdown("### 2. Análisis de Concentración (Pareto 80/20)")
            client_debt = df[df['dias'] > 0].groupby('cliente')['saldo'].sum().sort_values(ascending=False)
            
            if not client_debt.empty:
                client_debt_cumsum = client_debt.cumsum()
                total_debt_vencida = client_debt.sum()
                pareto_limit = total_debt_vencida * 0.80
                
                # Encontrar el punto de corte del 80%
                pareto_clients_df = client_debt.to_frame().iloc[0:len(client_debt_cumsum[client_debt_cumsum <= pareto_limit]) + 1]
                num_total_clientes_deuda = df[df['dias'] > 0]['cliente'].nunique()
                num_clientes_pareto = len(pareto_clients_df)
                porcentaje_clientes_pareto = (num_clientes_pareto / num_total_clientes_deuda) * 100 if num_total_clientes_deuda > 0 else 0
                
                st.info(f"Solo el **{porcentaje_clientes_pareto:.0f}%** de los clientes ({num_clientes_pareto} de {num_total_clientes_deuda}) representan aprox. el **80%** de la cartera vencida. **Priorice la gestión en estos {num_clientes_pareto} clientes.**")
                
                # Crear gráfico de Pareto (Combinación de Barra y Línea Acumulada)
                df_pareto_chart = client_debt.head(15).reset_index()
                df_pareto_chart['Acumulado'] = (df_pareto_chart['saldo'].cumsum() / total_debt_vencida)
                
                fig_pareto = go.Figure()
                fig_pareto.add_trace(go.Bar(x=df_pareto_chart['cliente'], y=df_pareto_chart['saldo'], name='Monto Vencido', marker_color=COLOR_PRIMARIO))
                fig_pareto.add_trace(go.Scatter(x=df_pareto_chart['cliente'], y=df_pareto_chart['Acumulado'], name='Acumulado', yaxis='y2', marker_color='#DC3545', line=dict(width=3)))
                fig_pareto.update_layout(
                    title='Top 15 Clientes por Deuda Vencida',
                    yaxis=dict(title='Monto Vencido', tickformat='$,.0f'),
                    yaxis2=dict(title='Porcentaje Acumulado', overlaying='y', side='right', tickformat='.0%')
                )
                st.plotly_chart(fig_pareto, use_container_width=True)

        # Análisis por Vendedor
        if sel_vendedor == "TODOS" and not df.empty:
            st.markdown("---")
            st.subheader("Desempeño por Vendedor/Zona")
            df_vendedor_resumen = df.groupby('vendedor').agg(
                Total_Cartera=('saldo', 'sum'),
                Vencido=('saldo', lambda x: df.loc[x.index][df.loc[x.index, 'dias'] > 0]['saldo'].sum()),
                Clientes=('cliente', 'nunique')
            ).reset_index()
            
            df_vendedor_resumen['% Vencido'] = (df_vendedor_resumen['Vencido'] / df_vendedor_resumen['Total_Cartera'] * 100).fillna(0)
            df_vendedor_resumen = df_vendedor_resumen.sort_values('% Vencido', ascending=False)
            
            st.dataframe(
                df_vendedor_resumen.style.format(
                    {'Total_Cartera': '${:,.0f}', 'Vencido': '${:,.0f}', '% Vencido': '{:.1f}%'}
                ).background_gradient(cmap='YlOrRd', subset=['% Vencido']),
                use_container_width=True, 
                hide_index=True
            )


    # --------------------------------------------------------
    # TAB 3: EXPORTACIÓN (Datos y Descargas)
    # --------------------------------------------------------
    with tab_export:
        st.subheader("📥 Descarga de Reportes y Detalle de Datos")
        
        col_dl, col_raw = st.columns([1, 2])
        
        with col_dl:
            st.markdown("**Reporte Listo para Gerencia**")
            excel_data = generar_excel_gerencial(df, kpis)
            st.download_button(
                label="✅ DESCARGAR EXCEL GERENCIAL FORMATEADO",
                data=excel_data,
                file_name=f"Reporte_Cartera_Ferreinox_{datetime.now().strftime('%Y-%m-%d')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                type="primary"
            )
            st.caption("Incluye KPIs, Resumen Ejecutivo y formato profesional con filtros y colores.")
            
        with col_raw:
            st.markdown("**Vista de la Base de Datos Filtrada**")
            # Mostrar todas las columnas relevantes
            st.dataframe(df.drop(columns=['Prioridad'], errors='ignore'), use_container_width=True, height=300)

if __name__ == "__main__":
    main()
