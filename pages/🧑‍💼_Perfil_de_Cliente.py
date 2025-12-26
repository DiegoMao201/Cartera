# ======================================================================================
# ARCHIVO: Tablero_Comando_Ferreinox_PRO.py (v.FINAL CORREGIDO - EMPLEADOS + EXCEL OK)
# Descripción: Panel de Control de Cartera PRO.
#              - Corrección: Error NameError en antiguedad_prom_vencida solucionado.
#              - Reporte Excel Gerencial de Solo Mora en Tab 1.
#              - Pestaña "Empleados" (Cruce con Excel Dropbox + Msj Nómina)
# ======================================================================================
import streamlit as st
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
import io
import os
import re
import unicodedata
from datetime import datetime
from urllib.parse import quote
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.drawing.image import Image as XLImage
from fpdf import FPDF
import yagmail
import tempfile
import glob
from io import BytesIO, StringIO
import dropbox # Conexión a Dropbox
import toml # Para manejo de secretos

# --- 1. CONFIGURACIÓN DE PÁGINA Y COLORES INSTITUCIONALES ---

st.set_page_config(
    page_title="🛡️ Centro de Mando: Cobranza Ferreinox PRO",
    page_icon="🛡️",
    layout="wide",
    initial_sidebar_state="expanded"
)

# Paleta Institucional Solicitada
COLOR_PRIMARIO = "#B21917"       # Rojo Intenso (Títulos, Barras principales)
COLOR_SECUNDARIO = "#E73537"     # Rojo Claro (Alertas, Botones secundarios)
COLOR_TERCIARIO = "#F0833A"      # Naranja (Énfasis, Subtítulos)
COLOR_ACCION = "#F9B016"         # Amarillo/Dorado (Botones de acción, KPIs importantes)
COLOR_FONDO_CLARO = "#FEF4C0"    # Crema (Fondos suaves, Tablas)
COLOR_BLANCO = "#FFFFFF"
COLOR_NEGRO = "#000000"

# --- CSS PERSONALIZADO (Tipografía Quicksand) ---
st.markdown(f"""
<style>
    /* Importar fuente Quicksand de Google Fonts */
    @import url('https://fonts.googleapis.com/css2?family=Quicksand:wght@400;500;600;700&display=swap');

    /* Aplicar fuente a toda la aplicación */
    html, body, [class*="css"] {{
        font-family: 'Quicksand', sans-serif;
    }}

    .stApp {{ background-color: #f8f9fa; }} /* Un gris muy suave para no cansar la vista */

    /* Métricas: Tarjetas con sombra y borde institucional */
    .stMetric {{ 
        background-color: {COLOR_BLANCO}; 
        padding: 15px; 
        border-radius: 12px; 
        border-left: 6px solid {COLOR_PRIMARIO}; 
        box-shadow: 0 4px 12px rgba(0,0,0,0.08);
    }}
    
    /* Títulos */
    h1, h2, h3 {{ color: {COLOR_PRIMARIO} !important; font-weight: 700; }}
    h1 {{ border-bottom: 3px solid {COLOR_ACCION}; padding-bottom: 10px; }}
    
    /* Tabs */
    .stTabs [data-baseweb="tab-list"] {{ gap: 10px; }}
    .stTabs [data-baseweb="tab"] {{
        border-radius: 4px 4px 0 0;
        color: {COLOR_NEGRO};
        font-weight: 600;
    }}
    .stTabs [aria-selected="true"] {{
        background-color: {COLOR_FONDO_CLARO};
        border-bottom: 3px solid {COLOR_PRIMARIO};
        color: {COLOR_PRIMARIO};
        font-weight: bold;
    }}

    /* Botones */
    div.stButton > button:first-child {{
        background-color: {COLOR_ACCION};
        color: {COLOR_NEGRO};
        font-weight: bold;
        border: none;
        border-radius: 8px;
        box-shadow: 0 2px 4px rgba(0,0,0,0.1);
    }}
    div.stButton > button:hover {{
        background-color: {COLOR_TERCIARIO};
        color: {COLOR_BLANCO};
    }}

    /* Botón WhatsApp */
    a.wa-link {{
        text-decoration: none; display: block; padding: 12px; margin-top: 10px;
        background-color: #25D366; color: white; border-radius: 8px; font-weight: bold;
        text-align: center; box-shadow: 0 2px 4px rgba(0,0,0,0.2);
        transition: transform 0.2s;
    }}
    a.wa-link:hover {{ background-color: #128C7E; transform: scale(1.02); }}

    /* Inputs */
    div[data-baseweb="input"], div[data-baseweb="select"], div.st-multiselect {{
        background-color: {COLOR_BLANCO};
        border: 1px solid {COLOR_TERCIARIO};
        border-radius: 8px;
    }}
    
    /* Expander */
    .streamlit-expanderHeader {{
        font-family: 'Quicksand', sans-serif;
        font-weight: 600;
        color: {COLOR_PRIMARIO};
    }}
</style>
""", unsafe_allow_html=True)


# ======================================================================================
# 2. MOTOR DE CONEXIÓN, LIMPIEZA Y PROCESAMIENTO
# ======================================================================================

def hex_to_rgb(hex_color):
    """Convierte HEX a tupla RGB (para FPDF)."""
    hex_color = hex_color.lstrip('#')
    return tuple(int(hex_color[i:i+2], 16) for i in (0, 2, 4))

def normalizar_texto(texto):
    if not isinstance(texto, str): return str(texto)
    # Normaliza, quita tildes, pone en mayúsculas y quita caracteres especiales
    texto = unicodedata.normalize('NFD', texto).encode('ascii', 'ignore').decode("utf-8").upper().strip()
    return re.sub(r'[^\w\s\.]', '', texto).strip()

def normalizar_nombre(nombre: str) -> str:
    """Función para normalizar nombres de vendedores/clientes."""
    if not isinstance(nombre, str): return ""
    nombre = nombre.upper().strip().replace('.', '')
    nombre = ''.join(c for c in unicodedata.normalize('NFD', nombre) if unicodedata.category(c) != 'Mn')
    return ' '.join(nombre.split())

def limpiar_nit(valor):
    """Deja solo números para cruce de llaves primarias."""
    if pd.isna(valor): return ""
    return re.sub(r'\D', '', str(valor))

def procesar_dataframe_robusto(df_raw):
    """
    Procesa el DataFrame crudo leído de Dropbox (Cartera).
    """
    df = df_raw.copy()

    # 1. Renombrar columnas a snake_case
    df.columns = [normalizar_texto(c).lower().replace(' ', '_') for c in df.columns]

    # 2. Limpieza de Tipos de Datos
    df['nomvendedor'] = df['nomvendedor'].astype(str).str.strip()
    df['importe'] = pd.to_numeric(df['importe'], errors='coerce').fillna(0)
    
    # Manejo de Notas Crédito
    df['numero'] = pd.to_numeric(df['numero'], errors='coerce').fillna(0)
    df.loc[df['numero'] < 0, 'importe'] *= -1
    
    df['dias_vencido'] = pd.to_numeric(df['dias_vencido'], errors='coerce').fillna(0).astype(int)

    # Fechas
    if 'fecha_documento' in df.columns:
        df['fecha_documento'] = pd.to_datetime(df['fecha_documento'], errors='coerce')
    if 'fecha_vencimiento' in df.columns:
        df['fecha_vencimiento'] = pd.to_datetime(df['fecha_vencimiento'], errors='coerce')

    # Normalización de Vendedor
    df['nomvendedor_norm'] = df['nomvendedor'].apply(normalizar_nombre)
    
    # Normalizar NIT para cruces
    df['nit_clean'] = df['nit'].apply(limpiar_nit)

    # 3. Asignación de Zonas
    ZONAS_SERIE = { "PEREIRA": [155, 189, 158, 439], "MANIZALES": [157, 238], "ARMENIA": [156] }
    ZONAS_SERIE_STR = {zona: [str(s) for s in series] for zona, series in ZONAS_SERIE.items()}
    
    def asignar_zona_robusta(valor_serie):
        if pd.isna(valor_serie): return "OTRAS ZONAS"
        numeros_en_celda = re.findall(r'\d+', str(valor_serie))
        if not numeros_en_celda: return "OTRAS ZONAS"
        for zona, series_clave_str in ZONAS_SERIE_STR.items():
            if set(numeros_en_celda) & set(series_clave_str): return zona
        return "OTRAS ZONAS"
    
    df['zona'] = df['serie'].apply(asignar_zona_robusta)

    # 4. Filtrado de series basura (W, X)
    df['serie'] = df['serie'].astype(str)
    df = df[~df['serie'].str.contains('W|X', case=False, na=False)]

    # 5. Segmentación estratégica de cartera
    bins = [-float('inf'), 0, 15, 30, 60, 90, float('inf')]
    labels = ["🟢 Al Día", "🟡 Prev. (1-15)", "🟠 Riesgo (16-30)", "🔴 Crítico (31-60)", "🚨 Alto Riesgo (61-90)", "⚫ Legal (+90)"]
    df['Rango'] = pd.cut(df['dias_vencido'], bins=bins, labels=labels, right=True)

    # Limpieza final: Quitar saldos cero
    df = df[df['importe'] != 0].copy()

    return df

@st.cache_data(ttl=600) 
def cargar_datos_automaticos_dropbox():
    """
    Carga:
    1. Cartera Principal (.csv)
    2. Datos Empleados (.xlsx)
    Realiza el cruce y retorna el DF unificado.
    """
    try:
        # Credenciales
        APP_KEY = st.secrets["dropbox"]["app_key"]
        APP_SECRET = st.secrets["dropbox"]["app_secret"]
        REFRESH_TOKEN = st.secrets["dropbox"]["refresh_token"]

        with dropbox.Dropbox(app_key=APP_KEY, app_secret=APP_SECRET, oauth2_refresh_token=REFRESH_TOKEN) as dbx:
            
            # --- 1. CARGA CARTERA (CSV) ---
            path_cartera = '/data/cartera_detalle.csv'
            metadata_c, res_c = dbx.files_download(path=path_cartera)
            contenido_csv = res_c.content.decode('latin-1')
            
            nombres_cols = [
                'Serie', 'Numero', 'Fecha Documento', 'Fecha Vencimiento', 'Cod Cliente',
                'NombreCliente', 'Nit', 'Poblacion', 'Provincia', 'Telefono1', 'Telefono2',
                'NomVendedor', 'Entidad Autoriza', 'E-Mail', 'Importe', 'Descuento',
                'Cupo Aprobado', 'Dias Vencido'
            ]
            df_raw = pd.read_csv(StringIO(contenido_csv), header=None, names=nombres_cols, sep='|', engine='python')
            df_proc = procesar_dataframe_robusto(df_raw)

            # --- 2. CARGA EMPLEADOS (EXCEL) ---
            msg_empleados = ""
            try:
                path_empleados = '/data/datos_empleados.xlsx'
                metadata_e, res_e = dbx.files_download(path=path_empleados)
                
                # Leer Excel desde memoria
                with io.BytesIO(res_e.content) as excel_buffer:
                    df_empleados = pd.read_excel(excel_buffer)
                
                # Limpiar columnas empleados (se espera: NOMBRE, CEDULA, TELEFONO, CORREO)
                df_empleados.columns = [c.strip().upper() for c in df_empleados.columns]
                
                if 'CEDULA' in df_empleados.columns:
                    df_empleados['cedula_clean'] = df_empleados['CEDULA'].apply(limpiar_nit)
                    
                    # --- 3. CRUCE (MERGE) ---
                    # Hacemos Left Join: Cartera + Info Empleado
                    df_proc = df_proc.merge(
                        df_empleados[['CEDULA', 'NOMBRE', 'TELEFONO', 'CORREO', 'cedula_clean']], 
                        left_on='nit_clean', 
                        right_on='cedula_clean', 
                        how='left'
                    )
                    
                    # Flag para identificar empleados
                    df_proc['es_empleado'] = df_proc['cedula_clean'].notna()
                    
                    # Renombrar columnas de empleado para evitar confusión
                    df_proc.rename(columns={
                        'TELEFONO': 'tel_empleado',
                        'CORREO': 'email_empleado',
                        'NOMBRE': 'nombre_empleado_db'
                    }, inplace=True)
                    
                    msg_empleados = " + 👷 Empleados vinculados"
                else:
                    df_proc['es_empleado'] = False
                    msg_empleados = " (⚠️ Archivo empleados sin col CEDULA)"

            except Exception as e_emp:
                df_proc['es_empleado'] = False
                msg_empleados = f" (⚠️ No se cargó empleados: {str(e_emp)})"

            
        return df_proc, f"Conectado: **Dropbox ({metadata_c.name})**{msg_empleados}"
            
    except toml.TomlDecodeError:
        return None, "Error: Credenciales no configuradas en secrets.toml"
    except KeyError as ke:
        return None, f"Error: Clave faltante en secrets.toml: {ke}"
    except Exception as e:
        return None, f"Error al cargar datos desde Dropbox: {e}"

# ======================================================================================
# 3. INTELIGENCIA DE NEGOCIO (ESTRATEGIA Y FUNCIONES)
# ======================================================================================

def calcular_kpis(df):
    """Calcula los KPIs principales de cobranza."""
    total = df['importe'].sum()
    vencido_df = df[df['dias_vencido'] > 0]
    vencido = vencido_df['importe'].sum()
    pct_vencido = (vencido / total * 100) if total else 0
    clientes_mora = vencido_df['nombrecliente'].nunique()
    
    # CSI (Collection Severity Index)
    csi = (vencido_df['importe'] * vencido_df['dias_vencido']).sum() / total if total > 0 else 0
    
    # Antigüedad Promedio Vencida
    antiguedad_prom_vencida = (vencido_df['importe'] * vencido_df['dias_vencido']).sum() / vencido if vencido > 0 else 0
    
    return total, vencido, pct_vencido, clientes_mora, csi, antiguedad_prom_vencida

def generar_analisis_cartera(kpis: dict):
    """Genera comentarios de análisis IA basados en KPIs."""
    comentarios = []
    
    # Estilos de color para HTML
    c_critico = COLOR_PRIMARIO
    c_riesgo = COLOR_TERCIARIO
    c_ok = "green"
    
    # 1. Análisis de % Vencido
    if kpis['porcentaje_vencido'] > 30: 
        comentarios.append(f"<li style='color:{c_critico}'><b>Alerta Crítica (%):</b> El <b>{kpis['porcentaje_vencido']:.1f}%</b> de la cartera está vencida. Requiere acciones urgentes.</li>")
    elif kpis['porcentaje_vencido'] > 15: 
        comentarios.append(f"<li style='color:{c_riesgo}'><b>Advertencia (%):</b> Con un <b>{kpis['porcentaje_vencido']:.1f}%</b> de cartera vencida, intensificar gestión.</li>")
    else: 
        comentarios.append(f"<li style='color:{c_ok}'><b>Saludable (%):</b> Nivel eficiente (<b>{kpis['porcentaje_vencido']:.1f}%</b>).</li>")
        
    # 2. Análisis de Antigüedad
    if kpis['antiguedad_prom_vencida'] > 60: 
        comentarios.append(f"<li style='color:{c_critico}'><b>Riesgo Alto (Días):</b> Antigüedad promedio de <b>{kpis['antiguedad_prom_vencida']:.0f} días</b>. Deuda envejecida.</li>")
    elif kpis['antiguedad_prom_vencida'] > 30: 
        comentarios.append(f"<li style='color:{c_riesgo}'><b>Atención (Días):</b> Antigüedad de <b>{kpis['antiguedad_prom_vencida']:.0f} días</b>. Concentrar en rango 31-60.</li>")
    else:
        comentarios.append(f"<li style='color:{c_ok}'><b>Gestión Preventiva:</b> Antigüedad baja (<b>{kpis['antiguedad_prom_vencida']:.0f} días</b>).</li>")
        
    # 3. Análisis de CSI
    if kpis['csi'] > 15: 
        comentarios.append(f"<li style='color:{c_critico}'><b>Severidad Crítica (CSI {kpis['csi']:.1f}):</b> Impacto muy alto en flujo de caja.</li>")
    else: 
        comentarios.append(f"<li><b>Severidad Controlada:</b> CSI de {kpis['csi']:.1f}.</li>")
        
    return "<ul>" + "".join(comentarios) + "</ul>"

def generar_link_wa(telefono, cliente, saldo_vencido, dias_max, nit, cod_cliente):
    """Genera el link de WhatsApp con mensaje pre-cargado CLIENTES."""
    tel = re.sub(r'\D', '', str(telefono))
    # Ajuste para números Colombia sin el 57 o con el 3 inicial
    if len(tel) == 10 and tel.startswith('3'): tel = '57' + tel 
    if len(tel) < 10: return None
    
    cliente_corto = str(cliente).split()[0].title()
    portal_link = "https://ferreinoxtiendapintuco.epayco.me/recaudo/ferreinoxrecaudoenlinea/"
    
    if saldo_vencido <= 0:
        msg = (
            f"👋 Hola {cliente_corto}, te saludamos de Ferreinox.\n"
            f"¡Tu cuenta está al día! 🌟 Gracias por tu puntualidad.\n"
            f"Te enviamos tu estado de cuenta al correo."
        )
    elif dias_max <= 30:
        msg = (
            f"👋 Hola {cliente_corto}, de Ferreinox.\n"
            f"Recordatorio: Saldo vencido de *${saldo_vencido:,.0f}*.\n"
            f"Factura más antigua: *{dias_max} días*.\n\n"
            f"Paga aquí 🔗 {portal_link}\n"
            f"NIT: {nit} | Cód: {cod_cliente}\n\n"
            f"¡Gracias!"
        )
    else:
        msg = (
            f"🚨 URGENTE {cliente_corto}: Saldo de *${saldo_vencido:,.0f}* con *{dias_max} días* de mora en Ferreinox.\n"
            f"Requerimos su pago inmediato para evitar bloqueos.\n\n"
            f"Pague aquí 🔗 {portal_link}\n"
            f"NIT: {nit} | Cód: {cod_cliente}\n\n"
            f"Por favor confirmar fecha de pago."
        )
            
    return f"https://wa.me/{tel}?text={quote(msg)}"

def generar_link_wa_empleado(telefono, nombre_empleado, saldo_total, monto_descuento):
    """Genera link WhatsApp específico para NÓMINA DE EMPLEADOS."""
    tel = re.sub(r'\D', '', str(telefono))
    if len(tel) == 10 and tel.startswith('3'): tel = '57' + tel
    if len(tel) < 10: return None
    
    nombre_corto = str(nombre_empleado).split()[0].title()
    
    msg = (
        f"👷 Hola {nombre_corto}, te informamos de Contabilidad Ferreinox.\n\n"
        f"Actualmente tienes facturas pendientes en cartera por un total de: *${saldo_total:,.0f}*.\n\n"
        f"⚠️ Se ha programado un descuento de nómina para esta quincena por valor de: *${monto_descuento:,.0f}*.\n\n"
        f"Si tienes dudas, por favor acércate a cartera. ¡Gracias!"
    )
    
    return f"https://wa.me/{tel}?text={quote(msg)}"

# ======================================================================================
# 4. GENERADORES (PDF Y EXCEL)
# ======================================================================================

class PDF(FPDF):
    """Clase personalizada para generar PDF con estilos de Ferreinox."""
    def header(self):
        # Usamos Helvetica que es estándar
        self.set_font('Helvetica', 'B', 12)
        
        # Color Primario para el PDF
        rgb_primario = hex_to_rgb(COLOR_PRIMARIO)
        self.set_text_color(*rgb_primario) 
        
        try:
             self.image("LOGO FERREINOX SAS BIC 2024.png", 10, 8, 80) 
        except: 
            self.cell(80, 10, 'FERREINOX SAS BIC', 0, 0, 'L')
            
        self.set_font('Helvetica', 'B', 18)
        self.set_text_color(0, 0, 0)
        self.cell(0, 10, 'ESTADO DE CUENTA', 0, 1, 'R')
        self.set_font('Helvetica', 'I', 9)
        self.cell(0, 10, f'Generado el: {datetime.now().strftime("%Y-%m-%d %H:%M")}', 0, 1, 'R')
        self.ln(5)

    def footer(self):
        self.set_y(-30)
        self.set_font('Helvetica', 'I', 8)
        self.set_text_color(128, 128, 128)
        
        portal_link = "https://ferreinoxtiendapintuco.epayco.me/recaudo/ferreinoxrecaudoenlinea/"
        self.set_font('Helvetica', 'B', 10)
        rgb_secundario = hex_to_rgb(COLOR_SECUNDARIO)
        self.set_text_color(*rgb_secundario)
        self.cell(0, 5, 'Portal de Pagos Ferreinox (Clic Aquí)', 0, 1, 'C', link=portal_link)
        self.set_font('Helvetica', 'I', 8)
        self.set_text_color(128, 128, 128)
        self.cell(0, 5, f'Página {self.page_no()}', 0, 0, 'C')

def crear_pdf(df_cliente, total_vencido_cliente):
    """Genera el PDF de estado de cuenta detallado."""
    pdf = PDF()
    pdf.set_auto_page_break(auto=True, margin=35)
    pdf.add_page()
    
    if df_cliente.empty:
        pdf.set_font('Helvetica', 'B', 12); pdf.cell(0, 10, 'Sin facturas.', 0, 1, 'C')
        return bytes(pdf.output())
        
    row = df_cliente.iloc[0]
    rgb_primario = hex_to_rgb(COLOR_PRIMARIO)
    
    # --- Datos Cliente ---
    pdf.set_font("Helvetica", 'B', 11); pdf.set_text_color(*rgb_primario)
    pdf.cell(40, 6, "Cliente:", 0, 0); pdf.set_font("Helvetica", '', 11); pdf.set_text_color(0); pdf.cell(0, 6, str(row['nombrecliente']), 0, 1)
    pdf.set_font("Helvetica", 'B', 11); pdf.set_text_color(*rgb_primario)
    pdf.cell(40, 6, "NIT:", 0, 0); pdf.set_font("Helvetica", '', 11); pdf.set_text_color(0); pdf.cell(0, 6, str(row['nit']), 0, 1)
    pdf.ln(5)
    
    # --- Tabla Headers ---
    rgb_fondo_crema = hex_to_rgb(COLOR_FONDO_CLARO)
    pdf.set_font('Helvetica', 'B', 10)
    pdf.set_fill_color(*rgb_primario)
    pdf.set_text_color(255, 255, 255)
    
    pdf.cell(25, 8, "Factura", 1, 0, 'C', 1)
    pdf.cell(25, 8, "Días Mora", 1, 0, 'C', 1)
    pdf.cell(35, 8, "Fecha Doc.", 1, 0, 'C', 1)
    pdf.cell(35, 8, "Fecha Venc.", 1, 0, 'C', 1)
    pdf.cell(40, 8, "Saldo", 1, 1, 'C', 1)
    
    pdf.ln()
    
    # --- Data ---
    pdf.set_font("Helvetica", '', 10)
    total_cartera = 0
    rgb_alerta = hex_to_rgb(COLOR_FONDO_CLARO) # Usamos el crema para filas vencidas
    
    for _, item in df_cliente.sort_values(by='dias_vencido', ascending=False).iterrows():
        total_cartera += item['importe']
        
        # Color filas vencidas
        if item['dias_vencido'] > 0:
            pdf.set_fill_color(*rgb_alerta)
            pdf.set_text_color(*rgb_primario) # Texto rojo
        else:
            pdf.set_fill_color(255, 255, 255)
            pdf.set_text_color(0, 0, 0)
            
        fecha_doc = item['fecha_documento'].strftime('%d/%m/%Y') if pd.notna(item['fecha_documento']) else '-'
        fecha_venc = item['fecha_vencimiento'].strftime('%d/%m/%Y') if pd.notna(item['fecha_vencimiento']) else '-'
        
        pdf.cell(25, 7, str(int(item['numero'])), 1, 0, 'C', 1)
        pdf.cell(25, 7, str(int(item['dias_vencido'])), 1, 0, 'C', 1)
        pdf.cell(35, 7, fecha_doc, 1, 0, 'C', 1)
        pdf.cell(35, 7, fecha_venc, 1, 0, 'C', 1)
        pdf.cell(40, 7, f"${item['importe']:,.0f}", 1, 1, 'R', 1)
        
    # --- Totales ---
    pdf.ln(2)
    pdf.set_text_color(0)
    pdf.set_font("Helvetica", 'B', 11)
    pdf.cell(120, 8, "TOTAL CARTERA", 1, 0, 'R')
    pdf.cell(40, 8, f"${total_cartera:,.0f}", 1, 1, 'R')

    if total_vencido_cliente > 0:
        pdf.ln()
        pdf.set_fill_color(*rgb_primario)
        pdf.set_text_color(255, 255, 255)
        pdf.cell(120, 8, 'TOTAL VENCIDO A PAGAR', 1, 0, 'R', 1)
        pdf.cell(40, 8, f"${total_vencido_cliente:,.0f}", 1, 1, 'R', 1)
            
    return bytes(pdf.output())

def crear_excel_gerencial(df, total, vencido, pct_mora, clientes_mora, csi, antiguedad_prom_vencida):
    """
    Genera el reporte ejecutivo en Excel (PARA EL TAB 3).
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Resumen Gerencial"
    
    # Preparar colores (QUITAR EL #)
    c_primario = COLOR_PRIMARIO.replace("#", "")
    c_secundario = COLOR_SECUNDARIO.replace("#", "")
    c_fondo_claro = COLOR_FONDO_CLARO.replace("#", "")
    c_blanco = "FFFFFF"
    
    # Fuentes Quicksand en Excel
    font_main = Font(name='Quicksand', size=11)
    font_header = Font(name='Quicksand', size=11, bold=True, color=c_blanco)
    font_title = Font(name='Quicksand', size=16, bold=True, color=c_primario)
    font_kpi_val = Font(name='Quicksand', size=12, bold=True, color=c_primario)
    
    fill_blue = PatternFill("solid", fgColor=c_primario)
    fill_kpi = PatternFill("solid", fgColor=c_fondo_claro)
    
    ws['A1'] = "REPORTE GERENCIAL DE CARTERA - FERREINOX"
    ws['A1'].font = font_title
    
    # KPIs
    kpi_labels = ["Total Cartera", "Total Vencido", "% Mora", "Clientes en Mora", "Antigüedad Prom.", "CSI"]
    kpi_values = [total, vencido, pct_mora / 100, clientes_mora, antiguedad_prom_vencida, csi]
    formats = ['$#,##0', '$#,##0', '0.0%', '0', '0.0', '0.0']
    
    for i, (lab, val, fmt) in enumerate(zip(kpi_labels, kpi_values, formats)):
        col_letter = get_column_letter(i+1)
        c_lab = ws.cell(row=3, column=i+1, value=lab)
        c_lab.font = font_header; c_lab.fill = fill_blue; c_lab.alignment = Alignment(horizontal='center')
        ws.column_dimensions[col_letter].width = 20
        
        c_val = ws.cell(row=4, column=i+1, value=val)
        c_val.number_format = fmt
        c_val.font = font_kpi_val; c_val.fill = fill_kpi; c_val.alignment = Alignment(horizontal='center')

    # Tabla Detalle
    ws['A6'] = "DETALLE COMPLETO (Filtrable)"
    ws['A6'].font = Font(name='Quicksand', size=12, bold=True, color=c_secundario)
    
    cols = ['nombrecliente', 'nit', 'numero', 'nomvendedor', 'cod_cliente', 'Rango', 'zona', 'dias_vencido', 'importe', 'telefono1', 'email']
    df_detalle = df[cols].sort_values(by='dias_vencido', ascending=False).reset_index(drop=True)

    # Headers
    for col_num, col_name in enumerate(cols, 1):
        c = ws.cell(row=7, column=col_num, value=col_name.upper().replace('_', ' '))
        c.fill = fill_blue
        c.font = font_header
        
    # Data
    for row_num, row_data in enumerate(df_detalle.values, 8):
        for col_num, val in enumerate(row_data, 1):
            c = ws.cell(row=row_num, column=col_num, value=val)
            c.font = font_main
            if col_num == 9: c.number_format = '$#,##0'
            
    # Filtros
    ws.auto_filter.ref = f"A7:{get_column_letter(len(cols))}{len(df_detalle)+7}"
    for i in range(1, len(cols) + 1):
        ws.column_dimensions[get_column_letter(i)].width = 22 if i != 1 else 35
        
    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()

def crear_excel_cobranza_vencida(df):
    """
    Genera un Excel conciso y gerencial solo con la cartera vencida para gestión.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Gestión Mora"
    
    # 1. Preparar Colores (Sin # para OpenPyXL)
    c_primario = COLOR_PRIMARIO.replace("#", "")      # Rojo
    c_fondo_suave = COLOR_FONDO_CLARO.replace("#", "") # Crema
    c_blanco = "FFFFFF"
    
    # 2. Definir Estilos
    font_titulo = Font(name='Quicksand', size=14, bold=True, color=c_primario)
    font_header = Font(name='Quicksand', size=11, bold=True, color=c_blanco)
    font_body = Font(name='Quicksand', size=10, color="000000")
    font_body_bold = Font(name='Quicksand', size=10, bold=True, color="000000")
    
    fill_header = PatternFill("solid", fgColor=c_primario)
    fill_zebra = PatternFill("solid", fgColor=c_fondo_suave)
    
    border_thin = Border(left=Side(style='thin', color="DDDDDD"), 
                           right=Side(style='thin', color="DDDDDD"), 
                           top=Side(style='thin', color="DDDDDD"), 
                           bottom=Side(style='thin', color="DDDDDD"))

    # 3. Filtrar Data (Solo Vencidos)
    df_vencidos = df[df['dias_vencido'] > 0].copy()
    
    # Seleccionar columnas clave y ordenar
    cols_export = ['nombrecliente', 'nit', 'telefono1', 'numero', 'fecha_vencimiento', 'dias_vencido', 'importe']
    df_export = df_vencidos[cols_export].sort_values(by=['nombrecliente', 'dias_vencido'], ascending=[True, False])
    
    # Renombrar para encabezados bonitos
    headers = ["CLIENTE", "NIT", "CONTACTO", "FACTURA #", "VENCIMIENTO", "DÍAS MORA", "SALDO PENDIENTE"]
    
    # 4. Construcción del Excel
    
    # Título del Reporte
    ws['A1'] = "REPORTE DE COBRANZA - CLIENTES EN MORA"
    ws['A1'].font = font_titulo
    ws['A2'] = f"Generado el: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws['A2'].font = Font(name='Quicksand', size=9, italic=True)
    
    # Encabezados de Tabla (Fila 4)
    start_row = 4
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=start_row, column=col_idx, value=header)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border_thin

    # Datos
    current_row = start_row + 1
    current_client = ""
    
    for _, row in df_export.iterrows():
        # Extracción de valores
        vals = [
            row['nombrecliente'],
            row['nit'],
            str(row['telefono1']),
            row['numero'],
            row['fecha_vencimiento'],
            row['dias_vencido'],
            row['importe']
        ]
        
        # Estilo de fila
        is_new_client = (vals[0] != current_client)
        current_client = vals[0]
        
        for col_idx, value in enumerate(vals, 1):
            cell = ws.cell(row=current_row, column=col_idx, value=value)
            cell.font = font_body
            cell.border = border_thin
            
            # Formateos específicos
            if col_idx == 1: # Cliente (Negrita)
                cell.font = font_body_bold
            
            if col_idx == 5 and isinstance(value, datetime): # Fecha
                 cell.number_format = 'DD/MM/YYYY'
            
            if col_idx == 7: # Importe
                cell.number_format = '$ #,##0'
                cell.font = font_body_bold # Destacar la deuda
            
            if col_idx == 6: # Días Mora (Centrar)
                cell.alignment = Alignment(horizontal='center')

            if row['dias_vencido'] > 60:
                 pass

        current_row += 1

    # Ajuste de Ancho de Columnas
    widths = [40, 15, 15, 12, 15, 12, 20]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Filtros Automáticos
    ws.auto_filter.ref = f"A{start_row}:{get_column_letter(len(headers))}{current_row-1}"
    
    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


# ======================================================================================
# 5. CORREOS Y PLANTILLAS
# ======================================================================================
def enviar_correo(destinatario, asunto, cuerpo_html, pdf_bytes):
    """Envía correo usando yagmail."""
    tmp_path = ''
    try:
        email_user = st.secrets["email_credentials"]["sender_email"]
        email_pass = st.secrets["email_credentials"]["sender_password"]
    except KeyError:
        st.error("⚠️ Error config secretos email.")
        return False

    if not destinatario or '@' not in destinatario:
        st.error("⚠️ Correo inválido.")
        return False

    try:
        with tempfile.NamedTemporaryFile(delete=False, suffix=".pdf") as tmp:
            tmp.write(pdf_bytes)
            tmp_path = tmp.name

        with st.spinner(f"Enviando a {destinatario}..."):
            yag = yagmail.SMTP(email_user, email_pass)
            yag.send(to=destinatario, subject=asunto, contents=[cuerpo_html, tmp_path])
        
        os.remove(tmp_path)
        return True
    except Exception as e:
        st.error(f"Error enviando correo: {e}")
        if os.path.exists(tmp_path): os.remove(tmp_path)
        return False
        
# --- PLANTILLAS HTML ESTILO QUICKSAND ---

def plantilla_correo_vencido(cliente, saldo, dias, nit, cod_cliente, portal_link):
    dias_max_vencido = int(dias)
    return f"""
    <!doctype html>
    <html>
    <head>
    <style>
        @import url('https://fonts.googleapis.com/css2?family=Quicksand:wght@400;700&display=swap');
        body {{ font-family: 'Quicksand', sans-serif; background-color: #f4f4f4; color: #333; }}
        .card {{ background: #ffffff; padding: 40px; border-radius: 12px; max-width: 600px; margin: auto; border-top: 6px solid {COLOR_PRIMARIO}; box-shadow: 0 4px 10px rgba(0,0,0,0.1); }}
        h1 {{ color: {COLOR_PRIMARIO}; font-weight: 700; }}
        .btn {{ background-color: {COLOR_ACCION}; color: #000; padding: 14px 24px; text-decoration: none; border-radius: 8px; font-weight: bold; display: inline-block; margin-top: 20px; }}
        .alert-box {{ background-color: {COLOR_FONDO_CLARO}; color: {COLOR_PRIMARIO}; padding: 15px; border-radius: 8px; text-align: center; font-size: 1.2em; font-weight: bold; margin: 20px 0; }}
        .small {{ font-size: 0.9em; color: #666; }}
    </style>
    </head>
    <body>
        <div class="card">
            <h1>Hola, {cliente}</h1>
            <p>Te contactamos de <strong>Ferreinox SAS BIC</strong>. Hemos identificado un saldo pendiente en tu cuenta.</p>
            
            <div class="alert-box">
                Saldo Vencido: ${saldo:,.0f}<br>
                <span style="font-size:0.8em">Mora Máxima: {dias_max_vencido} días</span>
            </div>
            
            <p>Evita inconvenientes en tus despachos gestionando tu pago hoy.</p>
            
            <center>
                <a href="{portal_link}" class="btn">🚀 Pagar en Línea Ahora</a>
            </center>
            
            <br>
            <p class="small">NIT: {nit} | Código: {cod_cliente}</p>
            <hr style="border:0; border-top:1px solid #eee;">
            <p class="small">Si ya pagaste, por favor omite este mensaje.</p>
        </div>
    </body>
    </html>
    """

def plantilla_correo_al_dia(cliente, saldo_total):
    return f"""
    <!doctype html>
    <html>
    <head>
    <style>
        @import url('https://fonts.googleapis.com/css2?family=Quicksand:wght@400;700&display=swap');
        body {{ font-family: 'Quicksand', sans-serif; background-color: #f4f4f4; color: #333; }}
        .card {{ background: #ffffff; padding: 40px; border-radius: 12px; max-width: 600px; margin: auto; border-top: 6px solid #28a745; box-shadow: 0 4px 10px rgba(0,0,0,0.1); }}
        h1 {{ color: {COLOR_PRIMARIO}; font-weight: 700; }}
        .info-box {{ background-color: #e8f5e9; color: #155724; padding: 15px; border-radius: 8px; text-align: center; font-weight: bold; margin: 20px 0; }}
    </style>
    </head>
    <body>
        <div class="card">
            <h1>¡Gracias, {cliente}!</h1>
            <p>En <strong>Ferreinox SAS BIC</strong> valoramos tu cumplimiento.</p>
            
            <div class="info-box">
                🌟 ¡Tu cuenta está al día!
            </div>
            
            <p>Saldo corriente actual: <strong>${saldo_total:,.0f}</strong></p>
            <p>Adjuntamos tu estado de cuenta actualizado para tu control administrativo.</p>
            
            <br>
            <p style="font-size:0.9em; color:#666;">Cordialmente,<br>Equipo de Cartera Ferreinox</p>
        </div>
    </body>
    </html>
    """

# ======================================================================================
# 6. DASHBOARD PRINCIPAL (MAIN)
# ======================================================================================

def main():
    # --- AUTENTICACIÓN ---
    if 'authentication_status' not in st.session_state:
        st.session_state['authentication_status'] = False
        st.session_state['acceso_general'] = False
        st.session_state['vendedor_autenticado'] = None

    if not st.session_state['authentication_status']:
        st.markdown(f"<h1 style='text-align:center; color:{COLOR_PRIMARIO}'>🔐 Acceso Ferreinox PRO</h1>", unsafe_allow_html=True)
        
        try:
            general_password = st.secrets["general"]["password"]
            vendedores_secrets = st.secrets["vendedores"]
        except Exception as e:
            st.error(f"Error secretos: {e}")
            st.stop()

        password = st.text_input("Contraseña:", type="password")
        if st.button("Ingresar"):
            if password == str(general_password):
                st.session_state['authentication_status'] = True
                st.session_state['acceso_general'] = True
                st.session_state['vendedor_autenticado'] = "GERENCIA"
                st.rerun()
            else:
                for k, v in vendedores_secrets.items():
                    if password == str(v):
                        st.session_state['authentication_status'] = True
                        st.session_state['acceso_general'] = False
                        st.session_state['vendedor_autenticado'] = k
                        st.rerun()
                st.error("Acceso Denegado")
        st.stop()
        
    # --- SIDEBAR ---
    with st.sidebar:
        try:
            st.image("LOGO FERREINOX SAS BIC 2024.png", use_container_width=True) 
        except: pass
        
        st.header("👤 Usuario Activo")
        st.success(f"**{st.session_state['vendedor_autenticado']}**")
        
        if st.button("Cerrar Sesión"):
            st.session_state.clear()
            st.rerun()
            
        st.divider()
        if st.button("🔄 Recargar Dropbox", type="primary"):
            st.cache_data.clear()
            st.rerun()

    # --- CARGA DATOS ---
    df, status = cargar_datos_automaticos_dropbox()
    st.sidebar.caption(status)
    if df is None: st.stop()

    # --- FILTROS ---
    st.sidebar.header("🔍 Filtros")
    
    # Filtro Vendedor
    if st.session_state['acceso_general']:
        opts = ["TODOS"] + sorted(df['nomvendedor'].unique().tolist())
        sel_vend = st.sidebar.selectbox("Vendedor:", opts)
        df_view = df if sel_vend == "TODOS" else df[df['nomvendedor_norm'] == normalizar_nombre(sel_vend)]
    else:
        df_view = df[df['nomvendedor_norm'] == normalizar_nombre(st.session_state['vendedor_autenticado'])]
        st.sidebar.info("Vista Vendedor")

    # Filtro Rango
    rangos = ["TODOS"] + df['Rango'].cat.categories.tolist()
    sel_rango = st.sidebar.selectbox("Antigüedad:", rangos)
    if sel_rango != "TODOS": df_view = df_view[df_view['Rango'] == sel_rango]

    # Filtro Zona
    zonas = ["TODAS"] + sorted(df_view['zona'].unique().tolist())
    sel_zona = st.sidebar.selectbox("Zona:", zonas)
    if sel_zona != "TODAS": df_view = df_view[df_view['zona'] == sel_zona]

    if df_view.empty:
        st.warning("Sin datos para los filtros seleccionados.")
        st.stop()

    # --- MAIN UI ---
    st.title("🛡️ Centro de Mando: Cobranza PRO")
    
    # KPIs
    # CORRECCIÓN AQUÍ: Usamos el nombre completo de la variable para evitar NameError
    total, vencido, pct, cli_mora, csi, antiguedad_prom_vencida = calcular_kpis(df_view)
    
    c1, c2, c3, c4, c5 = st.columns(5)
    c1.metric("Cartera Total", f"${total:,.0f}")
    c2.metric("Vencido", f"${vencido:,.0f}")
    c3.metric("% Vencido", f"{pct:.1f}%")
    c4.metric("Clientes Mora", cli_mora)
    c5.metric("CSI (Severidad)", f"{csi:.1f}")
    
    with st.expander("🤖 Análisis IA", expanded=(pct > 15)):
        # Pasamos la variable correcta 'antiguedad_prom_vencida'
        st.markdown(generar_analisis_cartera({'porcentaje_vencido': pct, 'antiguedad_prom_vencida': antiguedad_prom_vencida, 'csi': csi}), unsafe_allow_html=True)
        
    st.divider()
    
    # TABS (AÑADIDA TAB 4: EMPLEADOS)
    tab1, tab2, tab3, tab4 = st.tabs(["📞 GESTIÓN 1 a 1", "📊 ESTRATEGIA", "📥 DATA & REPORTES", "👷 EMPLEADOS"])
    
    # --- TAB 1: GESTIÓN ---
    with tab1:
        # ---- BOTÓN DE EXCEL GERENCIAL SOLO MORA ----
        col_header_1, col_header_2 = st.columns([3, 1])
        with col_header_1:
            st.subheader("🎯 Gestión de Cobro Directo")
        with col_header_2:
            excel_mora_bytes = crear_excel_cobranza_vencida(df_view)
            st.download_button(
                label="💾 Descargar Listado Mora (Excel)",
                data=excel_mora_bytes,
                file_name=f"Gestion_Mora_{datetime.now().strftime('%Y%m%d')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="btn_descarga_mora"
            )
        # --------------------------------------------------

        df_g = df_view[df_view['importe'] > 0].copy()
        
        if df_g.empty:
            st.info("No hay saldos pendientes para gestionar.")
        else:
            df_g['importe_vencido'] = df_g['importe'].where(df_g['dias_vencido'] > 0, 0)
            
            grp = df_g.groupby('nombrecliente').agg(
                saldo=('importe', 'sum'),
                vencido=('importe_vencido', 'sum'),
                dias_max=('dias_vencido', 'max'),
                tel=('telefono1', 'first'),
                email=('email', 'first'),
                nit=('nit', 'first'),
                cod=('cod_cliente', 'first')
            ).sort_values('vencido', ascending=False).reset_index()
            
            sel_cli = st.selectbox("Seleccionar Cliente (Ordenado por Vencido)", [""] + grp['nombrecliente'].tolist())
            
            if sel_cli:
                dat = grp[grp['nombrecliente'] == sel_cli].iloc[0]
                dets = df_view[df_view['nombrecliente'] == sel_cli].sort_values('dias_vencido', ascending=False)
                
                colA, colB = st.columns([1, 2])
                
                with colA:
                    st.markdown(f"#### {sel_cli}")
                    st.info(f"Total: ${dat['saldo']:,.0f}")
                    
                    if dat['vencido'] > 0:
                        st.markdown(f"<div style='background:{COLOR_FONDO_CLARO}; padding:10px; border-radius:5px; color:{COLOR_PRIMARIO}; font-weight:bold'>Vencido: ${dat['vencido']:,.0f}</div>", unsafe_allow_html=True)
                        st.error(f"Máx Mora: {int(dat['dias_max'])} días")
                    else:
                        st.success("✅ Al Día")
                        
                    pdf_bytes = crear_pdf(dets, dat['vencido'])
                    st.download_button("📄 PDF Estado Cuenta", pdf_bytes, f"EC_{sel_cli}.pdf", "application/pdf")
                    
                    st.divider()
                    st.markdown("#### 💬 WhatsApp Directo")
                    st.caption("Verifica el número antes de enviar:")
                    
                    # --- LÓGICA DE WA EDITABLE ---
                    raw_tel = str(dat['tel']) if pd.notna(dat['tel']) else ""
                    raw_tel = re.sub(r'\D', '', raw_tel) # Solo dejar dígitos
                    
                    telefono_destino = st.text_input("📱 Celular (Editable):", value=raw_tel, max_chars=15, help="Puedes escribir cualquier número aquí.")
                    
                    if telefono_destino:
                        wa_link = generar_link_wa(telefono_destino, sel_cli, dat['vencido'], dat['dias_max'], dat['nit'], dat['cod'])
                        
                        if wa_link:
                            st.markdown(f"""
                                <a href='{wa_link}' target='_blank' class='wa-link'>
                                🚀 Enviar Mensaje a {telefono_destino}
                                </a>
                            """, unsafe_allow_html=True)
                        else:
                            st.warning("⚠️ El número ingresado no parece válido (muy corto).")
                    else:
                        st.info("ℹ️ Ingresa un número de celular para generar el enlace.")

                with colB:
                    st.dataframe(dets[['numero', 'dias_vencido', 'fecha_vencimiento', 'importe', 'Rango']].style.format({'importe':'${:,.0f}'}), hide_index=True)
                    
                    st.write("#### 📧 Enviar Correo")
                    with st.form("frm_mail"):
                        dest = st.text_input("Email", value=dat['email'])
                        sub_btn = st.form_submit_button("Enviar PDF")
                        
                        if sub_btn:
                            subj = f"Estado de Cuenta - {sel_cli}"
                            if dat['vencido'] > 0:
                                body = plantilla_correo_vencido(sel_cli, dat['vencido'], dat['dias_max'], dat['nit'], dat['cod'], "https://ferreinoxtiendapintuco.epayco.me/")
                            else:
                                body = plantilla_correo_al_dia(sel_cli, dat['saldo'])
                                
                            if enviar_correo(dest, subj, body, pdf_bytes):
                                st.success("✅ Enviado correctamente")

    # --- TAB 2: ESTRATEGIA ---
    with tab2:
        c_pie, c_bar = st.columns(2)
        with c_pie:
            st.write("**Distribución por Mora**")
            grp_pie = df_view.groupby('Rango', observed=True)['importe'].sum().reset_index()
            # Mapa de colores para Plotly
            color_map = {
                "🟢 Al Día": "green", 
                "🟡 Prev. (1-15)": COLOR_FONDO_CLARO, # Crema
                "🟠 Riesgo (16-30)": COLOR_ACCION,    # Amarillo
                "🔴 Crítico (31-60)": COLOR_TERCIARIO, # Naranja
                "🚨 Alto Riesgo (61-90)": COLOR_SECUNDARIO, # Rojo claro
                "⚫ Legal (+90)": COLOR_PRIMARIO # Rojo oscuro
            }
            fig1 = px.pie(grp_pie, names='Rango', values='importe', color='Rango', color_discrete_map=color_map, hole=0.4)
            st.plotly_chart(fig1, use_container_width=True)
            
        with c_bar:
            st.write("**Top 10 Morosos**")
            top10 = df_view[df_view['dias_vencido']>0].groupby('nombrecliente')['importe'].sum().nlargest(10).reset_index()
            fig2 = px.bar(top10, x='importe', y='nombrecliente', orientation='h', text_auto='$.2s', color_discrete_sequence=[COLOR_PRIMARIO])
            fig2.update_layout(yaxis={'categoryorder':'total ascending'})
            st.plotly_chart(fig2, use_container_width=True)

        st.markdown("---")
        st.write("### Desempeño por Vendedor")
        
        # Tabla resumen vendedor
        res_vend = df_view.groupby('nomvendedor_norm').agg(
            Vendedor=('nomvendedor', 'first'),
            Total=('importe', 'sum'),
            Vencido=('importe', lambda x: x[df_view.loc[x.index, 'dias_vencido'] > 0].sum())
        ).reset_index()
        
        res_vend['% Vencido'] = (res_vend['Vencido'] / res_vend['Total'] * 100).fillna(0)
        
        # CSI por vendedor
        def calc_csi_vend(v):
            d = df_view[(df_view['nomvendedor_norm'] == v) & (df_view['dias_vencido'] > 0)]
            if d.empty: return 0
            return (d['importe'] * d['dias_vencido']).sum() / df_view[df_view['nomvendedor_norm'] == v]['importe'].sum()

        res_vend['CSI'] = res_vend['nomvendedor_norm'].apply(calc_csi_vend)
        
        st.dataframe(res_vend[['Vendedor', 'Total', 'Vencido', '% Vencido', 'CSI']].style.format({
            'Total': '${:,.0f}', 'Vencido': '${:,.0f}', '% Vencido': '{:.1f}%', 'CSI': '{:.1f}'
        }).background_gradient(subset=['% Vencido'], cmap='OrRd'), use_container_width=True, hide_index=True)

    # --- TAB 3: DATA ---
    with tab3:
        st.subheader("📥 Exportación")
        # CORRECCIÓN: Ahora pasamos 'antiguedad_prom_vencida' correctamente definida
        excel_bytes = crear_excel_gerencial(df_view, total, vencido, pct, cli_mora, csi, ant_prom)
        
        st.download_button(
            "💾 Descargar Reporte Gerencial (Excel)", 
            excel_bytes, 
            f"Cartera_{datetime.now().strftime('%Y%m%d')}.xlsx",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            type="primary"
        )
        
        st.subheader("Datos Crudos")
        st.dataframe(df_view, height=300)

    # --- TAB 4: EMPLEADOS (NUEVA PESTAÑA) ---
    with tab4:
        st.markdown("## 👷 Gestión de Cobro a Empleados")
        st.markdown("Este módulo cruza la cartera activa con la base de datos de empleados para gestionar descuentos de nómina.")
        
        # Filtrar solo empleados que tengan saldo > 0
        df_emps = df[df['es_empleado'] == True].copy()
        
        if df_emps.empty:
            st.info("ℹ️ No se encontraron empleados con cartera pendiente en este momento.")
        else:
            # Agrupar por empleado
            grp_emps = df_emps.groupby('nombre_empleado_db').agg(
                Cedula=('cedula_clean', 'first'),
                Total_Deuda=('importe', 'sum'),
                Telefono=('tel_empleado', 'first'),
                Email=('email_empleado', 'first'),
                Cant_Facturas=('numero', 'count')
            ).reset_index()
            
            # Filtrar solo los que deben algo (saldo > 0)
            grp_emps = grp_emps[grp_emps['Total_Deuda'] > 0].sort_values('Total_Deuda', ascending=False)
            
            if grp_emps.empty:
                st.success("✅ ¡Excelente! Todos los empleados identificados están Paz y Salvo.")
            else:
                col_izq, col_der = st.columns([1, 2])
                
                with col_izq:
                    st.subheader("👤 Seleccionar Empleado")
                    sel_empleado = st.selectbox("Lista de Empleados con Deuda:", grp_emps['nombre_empleado_db'].tolist())
                    
                    if sel_empleado:
                        data_emp = grp_emps[grp_emps['nombre_empleado_db'] == sel_empleado].iloc[0]
                        
                        st.markdown(f"""
                        <div style="background-color:{COLOR_FONDO_CLARO}; padding:15px; border-radius:10px; border-left:5px solid {COLOR_PRIMARIO}">
                            <h3 style="margin:0; color:{COLOR_PRIMARIO}">{sel_empleado}</h3>
                            <p><strong>CC:</strong> {data_emp['Cedula']}</p>
                            <p><strong>Deuda Total:</strong> <span style="font-size:1.2em; font-weight:bold">${data_emp['Total_Deuda']:,.0f}</span></p>
                            <p><strong>Facturas Pendientes:</strong> {data_emp['Cant_Facturas']}</p>
                        </div>
                        """, unsafe_allow_html=True)
                        
                        st.divider()
                        st.markdown("### 💸 Programar Descuento")
                        
                        # Input para valor a descontar
                        monto_descuento = st.number_input(
                            "Valor a descontar en quincena ($):", 
                            min_value=0.0, 
                            max_value=float(data_emp['Total_Deuda']), 
                            value=float(data_emp['Total_Deuda']),
                            step=1000.0,
                            format="%.0f"
                        )
                        
                        # Generación de Link WhatsApp
                        tel_emp_raw = str(data_emp['Telefono']) if pd.notna(data_emp['Telefono']) else ""
                        tel_emp_clean = re.sub(r'\D', '', tel_emp_raw)
                        
                        telefono_final = st.text_input("Confirmar Celular:", value=tel_emp_clean)
                        
                        if st.button("Generar Mensaje Nómina", type="primary"):
                            if monto_descuento > 0 and telefono_final:
                                link_wa_emp = generar_link_wa_empleado(
                                    telefono_final, 
                                    sel_empleado, 
                                    data_emp['Total_Deuda'], 
                                    monto_descuento
                                )
                                if link_wa_emp:
                                    st.markdown(f"""
                                    <a href='{link_wa_emp}' target='_blank' class='wa-link'>
                                    📤 Enviar Notificación de Descuento
                                    </a>
                                    """, unsafe_allow_html=True)
                                else:
                                    st.error("Número de teléfono inválido.")
                            else:
                                st.warning("Verifica el monto y el teléfono.")

                with col_der:
                    st.subheader("📄 Detalle de Facturas")
                    detalles_emp = df_emps[df_emps['nombre_empleado_db'] == sel_empleado][
                        ['numero', 'fecha_documento', 'fecha_vencimiento', 'dias_vencido', 'importe']
                    ].sort_values('dias_vencido', ascending=False)
                    
                    st.dataframe(
                        detalles_emp.style.format({
                            'importe': '${:,.0f}', 
                            'fecha_documento': '{:%Y-%m-%d}',
                            'fecha_vencimiento': '{:%Y-%m-%d}'
                        }), 
                        use_container_width=True,
                        hide_index=True
                    )

if __name__ == "__main__":
    main()
