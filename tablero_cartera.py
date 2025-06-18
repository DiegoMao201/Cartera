# ======================================================================================
# --- LIBRERÍAS Y CONFIGURACIÓN INICIAL ---
# ======================================================================================
import streamlit as st
import pandas as pd
import toml
import os
from io import BytesIO
import plotly.express as px
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
from openpyxl.worksheet.table import Table, TableStyleInfo
import unicodedata
import re
from datetime import datetime
from fpdf import FPDF # <-- NUEVA LIBRERÍA PARA PDF

st.set_page_config(
    page_title="Tablero de Cartera Ferreinox",
    page_icon="📊",
    layout="wide"
)

# ======================================================================================
# --- FUNCIONES AUXILIARES ---
# ======================================================================================

# --- NUEVA FUNCIÓN: GENERADOR DE PDF DE ESTADO DE CUENTA ---
class PDF(FPDF):
    def header(self):
        try:
            self.image("LOGO FERREINOX SAS BIC 2024.png", 10, 8, 80)
        except FileNotFoundError:
            self.set_font('Arial', 'B', 12)
            self.cell(0, 10, 'Logo no encontrado', 0, 1, 'C')
        self.set_font('Arial', 'B', 15)
        self.cell(80)
        self.cell(30, 10, 'Estado de Cuenta', 0, 1, 'C')
        self.set_font('Arial', 'I', 10)
        self.cell(80)
        self.cell(30, 10, f'Generado el: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}', 0, 1, 'C')
        self.ln(20)

    def footer(self):
        self.set_y(-30)
        self.set_font('Arial', 'B', 12)
        self.cell(0, 10, 'Realiza tu pago de forma facil y segura aqui:', 0, 1, 'C')
        self.set_font('Arial', 'U', 12)
        self.set_text_color(0, 0, 255)
        link = "https://ferreinoxtiendapintuco.epayco.me/recaudo/ferreinoxrecaudoenlinea/"
        self.cell(0, 10, link, 0, 1, 'C', link=link)

def generar_pdf_estado_cuenta(datos_cliente: pd.DataFrame):
    """Crea un PDF con el estado de cuenta de un cliente específico."""
    pdf = PDF()
    pdf.add_page()
    
    # Asegurarse de que hay datos antes de intentar acceder a ellos
    if datos_cliente.empty:
        pdf.set_font('Arial', 'B', 12)
        pdf.cell(0, 10, 'No se encontraron facturas para este cliente.', 0, 1, 'C')
        # --- LÍNEA DE CORRECCIÓN DEFINITIVA ---
        # Forzamos la salida a ser del tipo 'bytes', que es lo que Streamlit espera.
        return bytes(pdf.output())

    info_cliente = datos_cliente.iloc[0]
    
    # Datos del cliente
    pdf.set_font('Arial', 'B', 12)
    pdf.cell(0, 10, f"Cliente: {info_cliente['nombrecliente']}", 0, 1, 'L')
    pdf.set_font('Arial', '', 12)
    cod_cliente_str = str(int(info_cliente['cod_cliente'])) if pd.notna(info_cliente['cod_cliente']) else "N/A"
    pdf.cell(0, 10, f"Codigo de Cliente: {cod_cliente_str}", 0, 1, 'L')
    pdf.ln(10)

    # Encabezados de la tabla
    pdf.set_font('Arial', 'B', 10)
    pdf.set_fill_color(220, 220, 220)
    pdf.cell(30, 10, 'Factura', 1, 0, 'C', 1)
    pdf.cell(40, 10, 'Fecha Factura', 1, 0, 'C', 1)
    pdf.cell(40, 10, 'Fecha Vencimiento', 1, 0, 'C', 1)
    pdf.cell(40, 10, 'Importe', 1, 1, 'C', 1)

    # Contenido de la tabla
    pdf.set_font('Arial', '', 10)
    total_importe = 0
    for _, row in datos_cliente.iterrows():
        total_importe += row['importe']
        numero_factura_str = str(int(row['numero'])) if pd.notna(row['numero']) else "N/A"
        pdf.cell(30, 10, numero_factura_str, 1, 0, 'C')
        pdf.cell(40, 10, row['fecha_documento'].strftime('%Y-%m-%d'), 1, 0, 'C')
        pdf.cell(40, 10, row['fecha_vencimiento'].strftime('%Y-%m-%d'), 1, 0, 'C')
        pdf.cell(40, 10, f"${row['importe']:,.0f}", 1, 1, 'R')
        
    # Fila del total
    pdf.set_font('Arial', 'B', 10)
    pdf.cell(110, 10, 'TOTAL ADEUDADO', 1, 0, 'R', 1)
    pdf.cell(40, 10, f"${total_importe:,.0f}", 1, 1, 'R', 1)
    
    # --- LÍNEA DE CORRECCIÓN DEFINITIVA ---
    # Forzamos la salida a ser del tipo 'bytes', que es lo que Streamlit espera.
    # Esto soluciona el error sin importar el formato exacto que devuelva la librería.
    return bytes(pdf.output())


# (Aquí van el resto de tus funciones auxiliares: normalizar_nombre, procesar_cartera, etc.
# No las repito para brevedad, pero deben estar en tu script final)
def normalizar_nombre(nombre: str) -> str:
    if not isinstance(nombre, str): return ""
    nombre = nombre.upper().strip().replace('.', '')
    nombre = ''.join(c for c in unicodedata.normalize('NFD', nombre) if unicodedata.category(c) != 'Mn')
    return ' '.join(nombre.split())

ZONAS_SERIE = { "PEREIRA": [155, 189, 158, 439], "MANIZALES": [157, 238], "ARMENIA": [156] }

def procesar_cartera(df: pd.DataFrame) -> pd.DataFrame:
    df_proc = df.copy()
    df_proc['importe'] = pd.to_numeric(df_proc['importe'], errors='coerce').fillna(0)
    df_proc['dias_vencido'] = pd.to_numeric(df_proc['dias_vencido'], errors='coerce').fillna(0)
    df_proc['nomvendedor_norm'] = df_proc['nomvendedor'].apply(normalizar_nombre)
    ZONAS_SERIE_STR = {zona: [str(s) for s in series] for zona, series in ZONAS_SERIE.items()}
    def asignar_zona_robusta(valor_serie):
        if pd.isna(valor_serie): return "OTRAS ZONAS"
        numeros_en_celda = re.findall(r'\d+', str(valor_serie))
        if not numeros_en_celda: return "OTRAS ZONAS"
        for zona, series_clave_str in ZONAS_SERIE_STR.items():
            if set(numeros_en_celda) & set(series_clave_str): return zona
        return "OTRAS ZONAS"
    df_proc['zona'] = df_proc['serie'].apply(asignar_zona_robusta)
    bins = [-float('inf'), 0, 15, 30, 60, float('inf')]; labels = ['Al día', '1-15 días', '16-30 días', '31-60 días', 'Más de 60 días']
    df_proc['edad_cartera'] = pd.cut(df_proc['dias_vencido'], bins=bins, labels=labels, right=True)
    return df_proc

def generar_excel_formateado(df: pd.DataFrame):
    output = BytesIO()
    df_export = df[['nombrecliente', 'serie', 'numero', 'fecha_documento', 'fecha_vencimiento', 'importe', 'dias_vencido']].copy()
    for col in ['fecha_documento', 'fecha_vencimiento']: df_export[col] = pd.to_datetime(df_export[col], errors='coerce').dt.strftime('%d/%m/%Y')
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df_export.to_excel(writer, index=False, sheet_name='Cartera', startrow=9)
        wb, ws = writer.book, writer.sheets['Cartera']
        try:
            img = XLImage("LOGO FERREINOX SAS BIC 2024.png"); img.anchor = 'A1'; img.width = 390; img.height = 130
            ws.add_image(img)
        except FileNotFoundError: ws['A1'] = "Logo no encontrado."
        fill_red, fill_orange, fill_yellow = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid'), PatternFill(start_color='FFA500', end_color='FFA500', fill_type='solid'), PatternFill(start_color='FFF9C4', end_color='FFF9C4', fill_type='solid')
        font_bold, font_green_bold = Font(bold=True), Font(bold=True, color="006400")
        first_data_row, last_data_row = 10, ws.max_row
        tab = Table(displayName="CarteraVendedor", ref=f"A{first_data_row}:G{last_data_row}"); tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
        ws.add_table(tab)
        for i, ancho in enumerate([40, 10, 12, 18, 18, 18, 15], 1): ws.column_dimensions[get_column_letter(i)].width = ancho
        importe_col_idx, dias_col_idx, formato_moneda = 6, 7, '"$"#,##0'
        for row_idx, row in enumerate(ws.iter_rows(min_row=first_data_row, max_row=last_data_row), start=first_data_row):
            if row_idx == first_data_row:
                for cell in row: cell.font = font_bold; cell.alignment = Alignment(horizontal='center', vertical='center')
                continue
            row[importe_col_idx - 1].number_format = formato_moneda
            dias_cell = row[dias_col_idx - 1]
            dias = int(dias_cell.value) if str(dias_cell.value).isdigit() else 0
            if dias > 60: dias_cell.fill = fill_red
            elif dias > 30: dias_cell.fill = fill_orange
            elif dias > 0: dias_cell.fill = fill_yellow
            dias_cell.alignment = Alignment(horizontal='center')
        ws[f"E{last_data_row + 2}"] = "Tu cartera total es de:"; ws[f"E{last_data_row + 2}"].font = font_green_bold
        ws[f"F{last_data_row + 2}"] = f"=SUBTOTAL(9,F{first_data_row + 1}:F{last_data_row})"; ws[f"F{last_data_row + 2}"].number_format = formato_moneda; ws[f"F{last_data_row + 2}"].font = font_green_bold
        ws[f"E{last_data_row + 3}"] = "Facturas vencidas por valor de:"; ws[f"E{last_data_row + 3}"].font = font_green_bold
        ws[f"F{last_data_row + 3}"] = f"=SUMPRODUCT((SUBTOTAL(103,OFFSET(F{first_data_row+1},ROW(F{first_data_row+1}:F{last_data_row})-ROW(F{first_data_row+1}),0,1,1)))*(G{first_data_row+1}:G{last_data_row}>0),F{first_data_row+1}:F{last_data_row})"; ws[f"F{last_data_row + 3}"].number_format = formato_moneda; ws[f"F{last_data_row + 3}"].font = font_green_bold
    return output.getvalue()

@st.cache_data
def cargar_y_procesar_datos():
    cartera_df = pd.read_excel("Cartera.xlsx")
    cartera_df = cartera_df.iloc[:-1]
    cartera_df = cartera_df.rename(columns=lambda x: normalizar_nombre(x).lower().replace(' ', '_'))
    # Limpieza adicional para asegurar fechas correctas
    cartera_df['fecha_documento'] = pd.to_datetime(cartera_df['fecha_documento'], errors='coerce')
    cartera_df['fecha_vencimiento'] = pd.to_datetime(cartera_df['fecha_vencimiento'], errors='coerce')
    return procesar_cartera(cartera_df)

# ======================================================================================
# --- BLOQUE PRINCIPAL DE LA APP ---
# ======================================================================================
# (Autenticación, carga de datos y filtros - sin cambios)
try:
    general_password = st.secrets["general"]["password"]
    vendedores_secrets = st.secrets["vendedores"]
except Exception:
    st.error("Error al cargar las contraseñas desde los secretos.")
    st.info("Asegúrate de tener el archivo .streamlit/secrets.toml configurado correctamente si pruebas en local.")
    st.stop()

password = st.text_input("Introduce la contraseña para acceder a la cartera:", type="password")
if not password:
    st.warning("Debes ingresar una contraseña para continuar."); st.stop()

acceso_general, vendedor_autenticado = False, None
if password == str(general_password): acceso_general = True
else:
    for vendedor_key, pass_vendedor in vendedores_secrets.items():
        if password == str(pass_vendedor): vendedor_autenticado = vendedor_key; break
if not acceso_general and vendedor_autenticado is None:
    st.warning("Contraseña incorrecta. No tienes acceso al tablero."); st.stop()

st.title("📊 Tablero de Cartera Ferreinox SAS BIC")
try:
    cartera_procesada = cargar_y_procesar_datos()
except FileNotFoundError: st.error("No se encontró el archivo 'Cartera.xlsx'."); st.stop()
except Exception as e: st.error(f"Error al cargar o procesar 'Cartera.xlsx': {e}."); st.stop()

st.sidebar.title("Filtros")
vendedores_en_excel_display = sorted(cartera_procesada['nomvendedor'].dropna().unique())
if acceso_general:
    vendedor_sel = st.sidebar.selectbox("Filtrar por Vendedor:", ["Todos"] + vendedores_en_excel_display)
else:
    vendedor_autenticado_norm = normalizar_nombre(vendedor_autenticado)
    if vendedor_autenticado_norm not in cartera_procesada['nomvendedor_norm'].dropna().unique():
        st.error(f"¡Error de coincidencia! El vendedor '{vendedor_autenticado}' no se encontró en 'Cartera.xlsx'."); st.stop()
    vendedor_sel = vendedor_autenticado
    st.sidebar.success(f"Mostrando cartera de:"); st.sidebar.write(f"**{vendedor_sel}**")

lista_zonas = ["Todas las Zonas"] + list(ZONAS_SERIE.keys())
zona_sel = st.sidebar.selectbox("Filtrar por Zona:", lista_zonas)

if vendedor_sel == "Todos":
    cartera_filtrada = cartera_procesada.copy()
else:
    vendedor_sel_norm = normalizar_nombre(vendedor_sel)
    cartera_filtrada = cartera_procesada[cartera_procesada['nomvendedor_norm'] == vendedor_sel_norm].copy()
if zona_sel != "Todas las Zonas":
    cartera_filtrada = cartera_filtrada[cartera_filtrada['zona'] == zona_sel]

# (Renderizado de KPIs y Gráficos - sin cambios)
if cartera_filtrada.empty:
    st.warning(f"No se encontraron datos para la combinación de filtros seleccionada ('{vendedor_sel}' / '{zona_sel}')."); st.stop()

st.markdown("---")
total_cartera = cartera_filtrada['importe'].sum()
cartera_vencida = cartera_filtrada[cartera_filtrada['dias_vencido'] > 0]
total_vencido = cartera_vencida['importe'].sum()
porcentaje_vencido = (total_vencido / total_cartera) * 100 if total_cartera > 0 else 0
if total_cartera > 0:
    rotacion_dias = (cartera_filtrada['importe'] * cartera_filtrada['dias_vencido']).sum() / total_cartera
else: rotacion_dias = 0

if rotacion_dias <= 15: salud_rotacion, color_salud = "✅ Salud: Excelente", "green"
elif rotacion_dias <= 30: salud_rotacion, color_salud = "👍 Salud: Buena", "blue"
elif rotacion_dias <= 45: salud_rotacion, color_salud = "⚠️ Salud: Regular", "orange"
else: salud_rotacion, color_salud = "🚨 Salud: Alerta", "red"

col1, col2, col3, col4 = st.columns(4)
with col1: st.metric("💰 Cartera Total", f"${total_cartera:,.0f}")
with col2: st.metric("🔥 Cartera Vencida", f"${total_vencido:,.0f}")
with col3: st.metric("📈 % Vencido s/ Total", f"{porcentaje_vencido:.1f}%")
with col4:
    st.metric(label="🔄 Rotación (Días Promedio)", value=f"{rotacion_dias:.0f} días", help="Edad promedio de la cartera ponderada por el importe de cada factura.")
    st.markdown(f"<p style='color:{color_salud}; font-weight:bold; text-align:center;'>{salud_rotacion}</p>", unsafe_allow_html=True)

st.markdown("---")
col_grafico, col_tabla_resumen = st.columns([2, 1])
with col_grafico:
    st.subheader("Distribución de Cartera por Antigüedad")
    df_edades = cartera_filtrada.groupby('edad_cartera')['importe'].sum().reset_index()
    fig = px.bar(df_edades, x='edad_cartera', y='importe', text_auto='.2s', title='Monto de Cartera por Rango de Días', labels={'edad_cartera': 'Antigüedad', 'importe': 'Monto Total'}, color='edad_cartera', color_discrete_map={'Al día': 'green', '1-15 días': '#FFD700', '16-30 días': 'orange', '31-60 días': 'darkorange', 'Más de 60 días': 'red'})
    fig.update_layout(xaxis_title=None, yaxis_title="Monto ($)", showlegend=False)
    st.plotly_chart(fig, use_container_width=True)
with col_tabla_resumen:
    st.subheader("Resumen por Antigüedad")
    df_edades['Porcentaje'] = (df_edades['importe'] / total_cartera * 100).map('{:.1f}%'.format)
    df_edades['importe'] = df_edades['importe'].map('${:,.0f}'.format)
    st.dataframe(df_edades.rename(columns={'edad_cartera': 'Rango', 'importe': 'Monto'}), use_container_width=True, hide_index=True)

st.markdown("---")
# (Tabla de detalles principal - sin cambios)
st.subheader(f"Detalle de la Cartera: {vendedor_sel} / {zona_sel}")
st.download_button(label="📥 Descargar Reporte en Excel con Formato", data=generar_excel_formateado(cartera_filtrada), file_name=f'Cartera_{normalizar_nombre(vendedor_sel).replace(" ", "_")}_{zona_sel}.xlsx', mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
columnas_a_ocultar = ['provincia', 'telefono1', 'telefono2', 'entidad_autoriza', 'e_mail', 'descuento', 'cupo_aprobado', 'nomvendedor_norm', 'zona']
cartera_para_mostrar = cartera_filtrada.drop(columns=columnas_a_ocultar, errors='ignore')
st.dataframe(cartera_para_mostrar, use_container_width=True, hide_index=True)

# ======================================================================================
# --- NUEVA SECCIÓN: GENERADOR DE ESTADO DE CUENTA POR CLIENTE ---
# ======================================================================================
st.markdown("---")
st.header("⚙️ Herramientas de Gestión")
st.subheader("Generar Estado de Cuenta por Cliente")

# El buscador solo mostrará clientes que estén en la vista actualmente filtrada
lista_clientes = sorted(cartera_filtrada['nombrecliente'].dropna().unique())

if not lista_clientes:
    st.warning("No hay clientes para mostrar con los filtros actuales.")
else:
    cliente_seleccionado = st.selectbox(
        "Busca y selecciona un cliente para generar su estado de cuenta en PDF:",
        [""] + lista_clientes,
        format_func=lambda x: 'Selecciona un cliente...' if x == "" else x
    )

    if cliente_seleccionado:
        # Filtrar los datos para el cliente específico
        datos_cliente_seleccionado = cartera_filtrada[cartera_filtrada['nombrecliente'] == cliente_seleccionado].copy()
        
        st.write(f"**Facturas para {cliente_seleccionado}:**")
        st.dataframe(datos_cliente_seleccionado[['numero', 'fecha_documento', 'fecha_vencimiento', 'dias_vencido', 'importe']], use_container_width=True, hide_index=True)
        
        # Botón para generar y descargar el PDF
        st.download_button(
            label="📄 Descargar Estado de Cuenta (PDF)",
            data=generar_pdf_estado_cuenta(datos_cliente_seleccionado),
            file_name=f"Estado_Cuenta_{normalizar_nombre(cliente_seleccionado).replace(' ', '_')}.pdf",
            mime="application/pdf"
        )
